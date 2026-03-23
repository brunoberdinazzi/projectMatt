from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Tuple
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..models import (
    ChecklistDetail,
    ChecklistItem,
    ChecklistParseResult,
    FonteType,
    ParserOptions,
    ParserProfileDefinition,
    StatusType,
    WorkbookReferenceLink,
)
from .workbook_context_extractor import WorkbookContextExtractor

STATUS_MAP = {
    "sim": "Sim",
    "nao": "Nao",
    "não": "Nao",
    "n/a": "Nao se aplica",
    "na": "Nao se aplica",
    "parcialmente": "Parcialmente",
    "parcial": "Parcialmente",
    "parc": "Parcialmente",
}

RESPONSE_HEADER_RE = re.compile(r"resposta\s*\[(\d{4})\]", re.IGNORECASE)
DETAIL_COLUMN_PAIRS = (("D", "E"), ("G", "H"), ("J", "K"), ("M", "N"), ("P", "Q"))
DEFAULT_ALLOWED_GROUPS = ("1", "5")
DEFAULT_ALLOWED_STATUS = ("Nao", "Parcialmente")
AUTO_SHEET_SELECTIONS = {"", "*", "auto", "all", "todas", "todas_as_abas", "multi_aba"}
AUTO_PROFILE_SELECTIONS = {"", "auto", "automatico", "automatic", "detectar"}
FINANCIAL_MONTH_TITLES = {
    "janeiro",
    "fevereiro",
    "marco",
    "março",
    "abril",
    "maio",
    "junho",
    "julho",
    "agosto",
    "setembro",
    "outubro",
    "novembro",
    "dezembro",
}
NON_CRAWLABLE_EXTENSIONS = {
    ".pdf",
    ".csv",
    ".xls",
    ".xlsx",
    ".ods",
    ".doc",
    ".docx",
    ".odt",
    ".zip",
    ".rar",
    ".7z",
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
}
PARSER_PROFILE_MAP = {
    "auto": ParserProfileDefinition(
        key="auto",
        label="Automatico",
        description="Inspeciona a planilha e escolhe automaticamente entre checklist e financeiro / DRE antes da leitura.",
        allowed_groups=[],
        allowed_status=[],
    ),
    "default": ParserProfileDefinition(
        key="default",
        label="Padrao",
        description="Recorte padrao: grupos 1 e 5 com itens Nao ou Parcialmente.",
        allowed_groups=["1", "5"],
        allowed_status=["Nao", "Parcialmente"],
    ),
    "extended": ParserProfileDefinition(
        key="extended",
        label="Estendido",
        description="Checklist ampliado: grupos 1 a 5, mantendo foco em Nao ou Parcialmente.",
        allowed_groups=["1", "2", "3", "4", "5"],
        allowed_status=["Nao", "Parcialmente"],
    ),
    "full": ParserProfileDefinition(
        key="full",
        label="Completo",
        description="Checklist completo: grupos 1 a 5 e todos os status normalizados.",
        allowed_groups=["1", "2", "3", "4", "5"],
        allowed_status=["Sim", "Nao", "Parcialmente", "Nao se aplica"],
    ),
    "financial_dre": ParserProfileDefinition(
        key="financial_dre",
        label="Financeiro / DRE",
        description="Leitura de planilhas financeiras mensais para consolidacao de DRE, custos, recebiveis e resultado.",
        allowed_groups=[],
        allowed_status=[],
    ),
}


@dataclass
class ParserConfig:
    profile: str = "default"
    allowed_groups: set[str] = field(default_factory=lambda: set(DEFAULT_ALLOWED_GROUPS))
    allowed_status: set[str] = field(default_factory=lambda: set(DEFAULT_ALLOWED_STATUS))
    checklist_sheet_name: str = "auto"
    metadata_row: int = 5


class ChecklistParser:
    def __init__(self, config: ParserConfig) -> None:
        self.config = config
        self.context_extractor = WorkbookContextExtractor()

    def parse(self, workbook_path: Path, source_name: Optional[str] = None) -> ChecklistParseResult:
        result = ChecklistParseResult(
            grupos_permitidos=sorted(self.config.allowed_groups),
            parser_options=ParserOptions(
                profile=self.config.profile,
                allowed_groups=sorted(self.config.allowed_groups),
                allowed_status=sorted(self.config.allowed_status),
                checklist_sheet_name=self.config.checklist_sheet_name,
                metadata_row=self.config.metadata_row,
            ),
        )
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            sheets, missing_sheet_names = self._resolve_checklist_sheets(workbook)
            if not sheets:
                if missing_sheet_names:
                    result.warnings.append(
                        "Abas solicitadas nao encontradas: " + ", ".join(missing_sheet_names) + "."
                    )
                result.warnings.append("Nenhuma aba de checklist compativel foi encontrada.")
                return result
            result.parser_options.checklist_sheet_names = [sheet.title for sheet in sheets]

            filename_hint = Path(source_name) if source_name else workbook_path
            result.orgao = self._infer_orgao_from_filename(filename_hint)
            result.tipo_orgao = self._infer_tipo_orgao(filename_hint)
            result.sat_numero = self._infer_sat_numero(filename_hint)
            allowed_group_statuses: list[tuple[str, Optional[str]]] = []
            all_observations: list[tuple[str, str]] = []

            for sheet in sheets:
                self._merge_sheet_metadata(result, sheet)
                response_columns = self._detect_response_columns(sheet)
                self._merge_unique_values(
                    result.fontes_disponiveis,
                    self._detect_available_sources(sheet, response_columns),
                )
                observation_start_row = self._find_observation_start_row(sheet)
                observations = self._parse_observations(sheet, observation_start_row)
                all_observations.extend(observations)
                allowed_group_statuses.extend(
                    self._collect_allowed_group_statuses(sheet, response_columns, observation_start_row)
                )
                result.itens_processados.extend(
                    self._parse_items(
                        sheet=sheet,
                        response_columns=response_columns,
                        observation_start_row=observation_start_row,
                        observations=observations,
                        warnings=result.warnings,
                    )
                )

            if missing_sheet_names:
                result.warnings.append(
                    "Abas solicitadas nao encontradas: " + ", ".join(missing_sheet_names) + "."
                )

            result.context_layers = self.context_extractor.extract(workbook, result)
            result.reference_links = self._extract_reference_links(workbook, result)

            if not result.itens_processados:
                result.warnings.append("Nenhum item elegivel encontrado nas abas selecionadas.")
                self._append_empty_scope_warnings(
                    warnings=result.warnings,
                    observations=all_observations,
                    allowed_group_statuses=allowed_group_statuses,
                )

            return result
        finally:
            workbook.close()

    def _resolve_checklist_sheets(self, workbook) -> tuple[list, list[str]]:
        if _is_auto_sheet_selection(self.config.checklist_sheet_name):
            return self._discover_candidate_sheets(workbook), []
        requested_names = _parse_sheet_names(self.config.checklist_sheet_name)
        return self._resolve_named_sheets(workbook, requested_names)

    def _resolve_named_sheets(self, workbook, requested_names: list[str]) -> tuple[list, list[str]]:
        selected = []
        selected_normalized: set[str] = set()
        missing: list[str] = []

        for requested_name in requested_names:
            sheet = self._find_sheet_by_name(workbook, requested_name)
            if sheet is None:
                missing.append(requested_name)
                continue
            normalized_title = _normalize_text(sheet.title)
            if normalized_title in selected_normalized:
                continue
            selected.append(sheet)
            selected_normalized.add(normalized_title)

        return selected, missing

    def _discover_candidate_sheets(self, workbook) -> list:
        selected = []
        for sheet in workbook.worksheets:
            if self._looks_like_checklist_sheet(sheet):
                selected.append(sheet)
        return selected

    def _find_sheet_by_name(self, workbook, requested_name: str):
        if requested_name in workbook.sheetnames:
            return workbook[requested_name]

        requested_normalized = _normalize_text(requested_name)
        for sheet_name in workbook.sheetnames:
            if _normalize_text(sheet_name) == requested_normalized:
                return workbook[sheet_name]
        return None

    def _looks_like_checklist_sheet(self, sheet) -> bool:
        normalized_title = _normalize_text(sheet.title)
        if "checklist" in normalized_title:
            return True
        if (sheet.max_row or 0) < self.config.metadata_row:
            return False

        response_columns = self._detect_response_columns(sheet)
        if not response_columns:
            return False

        item_header = _normalize_text(sheet[f"B{self.config.metadata_row}"].value)
        description_header = _normalize_text(sheet[f"C{self.config.metadata_row}"].value)
        if item_header in {"item", "codigo", "item_codigo"} and description_header in {
            "descricao",
            "descricao_item",
            "descricao_do_item",
        }:
            return True

        return self._find_observation_start_row(sheet) <= (sheet.max_row or 0)

    def _merge_sheet_metadata(self, result: ChecklistParseResult, sheet) -> None:
        if not result.orgao:
            result.orgao = _clean_value(sheet["E3"].value) or result.orgao
        if not result.site_url:
            result.site_url = self._find_first_url_for_items(sheet, ["1.1"])
        if not result.portal_url:
            result.portal_url = self._find_first_url_for_items(sheet, ["1.2"])
        if not result.esic_url:
            result.esic_url = self._find_first_url_for_items(
                sheet,
                ["5.1", "5.2", "5.4", "5.5", "5.6", "5.7", "5.8"],
            )

    def _merge_unique_values(self, current_values: list[str], new_values: list[str]) -> None:
        for value in new_values:
            if value not in current_values:
                current_values.append(value)

    def _extract_reference_links(self, workbook, result: ChecklistParseResult) -> list[WorkbookReferenceLink]:
        links: list[WorkbookReferenceLink] = []
        seen_urls: set[str] = set()

        self._append_reference_link(
            links=links,
            seen_urls=seen_urls,
            url=result.site_url,
            sheet_name="Checklist",
            cell_reference=None,
            label="Canal principal identificado na leitura estruturada",
            context="URL principal usada como origem estruturada do crawler.",
            source_hint="site_orgao",
            link_kind="primary",
        )
        self._append_reference_link(
            links=links,
            seen_urls=seen_urls,
            url=result.portal_url,
            sheet_name="Checklist",
            cell_reference=None,
            label="Canal complementar identificado na leitura estruturada",
            context="URL complementar usada como origem estruturada do crawler.",
            source_hint="portal_transparencia",
            link_kind="primary",
        )
        self._append_reference_link(
            links=links,
            seen_urls=seen_urls,
            url=result.esic_url,
            sheet_name="Checklist",
            cell_reference=None,
            label="Canal de atendimento identificado na leitura estruturada",
            context="URL de atendimento usada como origem estruturada do crawler.",
            source_hint="esic",
            link_kind="primary",
        )

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                row_values = [
                    (cell.coordinate, _clean_value(cell.value))
                    for cell in row
                    if _clean_value(cell.value)
                ]
                if not row_values:
                    continue
                if not self._should_collect_reference_row(
                    sheet_name=sheet.title,
                    row_values=row_values,
                    selected_checklist_sheets=result.parser_options.checklist_sheet_names,
                    entity_name=result.orgao,
                ):
                    continue

                row_text = " | ".join(value for _, value in row_values)
                for coordinate, value in row_values:
                    for url in _extract_urls(value):
                        self._append_reference_link(
                            links=links,
                            seen_urls=seen_urls,
                            url=url,
                            sheet_name=sheet.title,
                            cell_reference=coordinate,
                            label=self._build_reference_label(row_values, coordinate, url),
                            context=_truncate_text(row_text, 240),
                            source_hint=self._infer_reference_source_hint(sheet.title, row_text, url),
                            link_kind="reference",
                        )

        return links

    def _append_reference_link(
        self,
        links: list[WorkbookReferenceLink],
        seen_urls: set[str],
        url: Optional[str],
        sheet_name: str,
        cell_reference: Optional[str],
        label: Optional[str],
        context: Optional[str],
        source_hint: FonteType,
        link_kind: str,
    ) -> None:
        if not url:
            return
        normalized_url = _normalize_url_for_dedup(url)
        if not normalized_url or normalized_url in seen_urls:
            return
        seen_urls.add(normalized_url)
        links.append(
            WorkbookReferenceLink(
                url=url,
                sheet_name=sheet_name,
                cell_reference=cell_reference,
                label=label,
                context=context,
                source_hint=source_hint,
                link_kind=link_kind,
                crawlable=_is_crawlable_reference_url(url),
            )
        )

    def _build_reference_label(
        self,
        row_values: list[tuple[str, str]],
        coordinate: str,
        url: str,
    ) -> Optional[str]:
        candidates = []
        for cell_coordinate, value in row_values:
            if cell_coordinate == coordinate:
                continue
            if url in value:
                continue
            candidates.append(value)
        if not candidates:
            return None
        return _truncate_text(candidates[0], 120)

    def _infer_reference_source_hint(self, sheet_name: str, context: str, url: str) -> FonteType:
        haystack = " ".join(part for part in [sheet_name, context, url] if part).lower()
        if "esic" in haystack or "e-sic" in haystack or "sic" in haystack:
            return "esic"
        if "portal" in haystack and "transpar" in haystack:
            return "portal_transparencia"
        if any(keyword in haystack for keyword in ("site", "prefeitura", "camara", "câmara", "orgao", "órgão")):
            return "site_orgao"
        return "nao_informada"

    def _should_collect_reference_row(
        self,
        sheet_name: str,
        row_values: list[tuple[str, str]],
        selected_checklist_sheets: list[str],
        entity_name: Optional[str],
    ) -> bool:
        normalized_sheet = _normalize_text(sheet_name)
        selected_sheets = {_normalize_text(name) for name in selected_checklist_sheets}
        if normalized_sheet in selected_sheets:
            return True
        if not entity_name:
            return False

        entity_key = _entity_key(entity_name)
        if not entity_key:
            return False

        for _, value in row_values:
            if _entity_key(value) == entity_key:
                return True
        return False

    def _detect_response_columns(self, sheet) -> list[tuple[str, str]]:
        response_columns: list[tuple[str, str]] = []
        if (sheet.max_row or 0) < self.config.metadata_row:
            return response_columns

        row_iter = sheet.iter_rows(min_row=self.config.metadata_row, max_row=self.config.metadata_row)
        metadata_cells = next(row_iter, [])
        for cell in metadata_cells:
            value = _clean_value(cell.value)
            if not value:
                continue
            match = RESPONSE_HEADER_RE.search(value)
            if match:
                response_columns.append((get_column_letter(cell.column), match.group(1)))

        return response_columns

    def _find_observation_start_row(self, sheet) -> int:
        for row_idx in range(1, (sheet.max_row or 0) + 1):
            label = _clean_value(sheet[f"B{row_idx}"].value)
            if label and _normalize_text(label).startswith("observacoes"):
                return row_idx

        return (sheet.max_row or 0) + 1

    def _parse_observations(self, sheet, observation_start_row: int) -> list[tuple[str, str]]:
        observations: list[tuple[str, str]] = []
        for row_idx in range(observation_start_row + 1, (sheet.max_row or 0) + 1):
            item_expr = _clean_value(sheet[f"B{row_idx}"].value)
            comment = _clean_value(sheet[f"E{row_idx}"].value)
            if not item_expr and not comment:
                continue
            if _normalize_text(item_expr) == "item":
                continue
            if item_expr and comment:
                observations.append((item_expr, comment))

        return observations

    def _parse_items(
        self,
        sheet,
        response_columns: list[tuple[str, str]],
        observation_start_row: int,
        observations: list[tuple[str, str]],
        warnings: list[str],
    ) -> list[ChecklistItem]:
        parsed_items: list[ChecklistItem] = []
        row_idx = self.config.metadata_row + 1
        while row_idx < observation_start_row:
            item_code = _clean_value(sheet[f"B{row_idx}"].value)
            description = _clean_value(sheet[f"C{row_idx}"].value)
            if not item_code:
                row_idx += 1
                continue

            if not description:
                row_idx += 1
                continue

            group = self._resolve_group(item_code)
            if group not in self.config.allowed_groups:
                row_idx = self._skip_detail_rows(sheet, row_idx + 1, observation_start_row)
                continue

            selected_status = self._select_reference_status(sheet, row_idx, response_columns)
            if selected_status is None:
                row_idx = self._skip_detail_rows(sheet, row_idx + 1, observation_start_row)
                continue

            status_year, raw_status = selected_status
            status = _normalize_status(raw_status)
            if status not in self.config.allowed_status:
                row_idx = self._skip_detail_rows(sheet, row_idx + 1, observation_start_row)
                continue

            details, next_row = self._collect_details(sheet, row_idx + 1, observation_start_row)
            observacao = self._find_observation(item_code, observations)
            fonte_texto = _clean_value(sheet[f"T{row_idx}"].value)
            fundamentacao = _clean_value(sheet[f"U{row_idx}"].value)

            parsed_items.append(
                ChecklistItem(
                    grupo=group,
                    item_codigo=item_code,
                    linha_referencia=row_idx,
                    ano_referencia=status_year,
                    status=status,
                    status_2024=_clean_value(sheet[f"R{row_idx}"].value),
                    status_2025=_clean_value(sheet[f"S{row_idx}"].value),
                    fonte=_normalize_fonte(fonte_texto),
                    fonte_texto=fonte_texto,
                    descricao_item=description,
                    observacao=observacao,
                    fundamentacao=fundamentacao,
                    detalhes=details,
                    aba_origem=sheet.title,
                )
            )

            if observacao is None:
                warnings.append(
                    f"Aba '{sheet.title}', linha {row_idx}: item {item_code} sem observacao vinculada."
                )

            row_idx = next_row

        return parsed_items

    def _collect_allowed_group_statuses(
        self,
        sheet,
        response_columns: list[tuple[str, str]],
        observation_start_row: int,
    ) -> list[tuple[str, Optional[str]]]:
        statuses: list[tuple[str, Optional[str]]] = []
        row_idx = self.config.metadata_row + 1
        while row_idx < observation_start_row:
            item_code = _clean_value(sheet[f"B{row_idx}"].value)
            description = _clean_value(sheet[f"C{row_idx}"].value)
            if not item_code or not description:
                row_idx += 1
                continue

            group = self._resolve_group(item_code)
            if group not in self.config.allowed_groups:
                row_idx = self._skip_detail_rows(sheet, row_idx + 1, observation_start_row)
                continue

            selected_status = self._select_reference_status(sheet, row_idx, response_columns)
            normalized_status = _normalize_status(selected_status[1]) if selected_status else None
            statuses.append((item_code, normalized_status))
            row_idx = self._skip_detail_rows(sheet, row_idx + 1, observation_start_row)

        return statuses

    def _append_empty_scope_warnings(
        self,
        warnings: list[str],
        observations: list[tuple[str, str]],
        allowed_group_statuses: list[tuple[str, Optional[str]]],
    ) -> None:
        if allowed_group_statuses and all(
            status not in self.config.allowed_status for _, status in allowed_group_statuses
        ):
            warnings.append(
                f"Nos grupos {', '.join(sorted(self.config.allowed_groups))}, os itens avaliados na planilha "
                f"estao fora do recorte automatizado de status {', '.join(sorted(self.config.allowed_status))}."
            )

        out_of_scope_observations = sorted(
            {
                expression
                for expression, _ in observations
                if self._resolve_group(expression) not in self.config.allowed_groups
            }
        )
        if out_of_scope_observations:
            warnings.append(
                "Foram encontradas observacoes apenas para itens fora do escopo automatizado atual "
                f"(grupos {', '.join(sorted(self.config.allowed_groups))}): {', '.join(out_of_scope_observations)}."
            )

    def _select_reference_status(
        self,
        sheet,
        row_idx: int,
        response_columns: list[tuple[str, str]],
    ) -> Optional[Tuple[str, str]]:
        for column_letter, year in reversed(response_columns):
            raw_value = _clean_value(sheet[f"{column_letter}{row_idx}"].value)
            normalized = _normalize_status(raw_value)
            if normalized is None or normalized == "Nao se aplica":
                continue
            return year, raw_value
        return None

    def _resolve_group(self, item_code: Optional[str]) -> Optional[str]:
        if not item_code:
            return None
        match = re.match(r"\s*(\d+)", item_code)
        if match:
            return match.group(1)
        return None

    def _collect_details(
        self,
        sheet,
        start_row: int,
        observation_start_row: int,
    ) -> tuple[list[ChecklistDetail], int]:
        details: list[ChecklistDetail] = []
        current_row = start_row
        while current_row < observation_start_row:
            if _clean_value(sheet[f"B{current_row}"].value):
                break

            row_details = self._parse_detail_row(sheet, current_row)
            if row_details:
                details.extend(row_details)

            current_row += 1

        return details, current_row

    def _parse_detail_row(self, sheet, row_idx: int) -> list[ChecklistDetail]:
        details: list[ChecklistDetail] = []
        for status_column, label_column in DETAIL_COLUMN_PAIRS:
            label = _clean_value(sheet[f"{label_column}{row_idx}"].value)
            status = _clean_value(sheet[f"{status_column}{row_idx}"].value)
            if not label:
                continue
            details.append(
                ChecklistDetail(
                    descricao=label,
                    status=_normalize_detail_status(status),
                )
            )
        return details

    def _skip_detail_rows(self, sheet, start_row: int, observation_start_row: int) -> int:
        current_row = start_row
        while current_row < observation_start_row and not _clean_value(sheet[f"B{current_row}"].value):
            current_row += 1
        return current_row

    def _find_observation(
        self,
        item_code: str,
        observations: list[tuple[str, str]],
    ) -> Optional[str]:
        matches = [
            comment
            for expression, comment in observations
            if _observation_expression_matches(expression, item_code)
        ]
        if not matches:
            return None
        return "\n".join(matches)

    def _infer_orgao_from_filename(self, workbook_path: Path) -> Optional[str]:
        stem = workbook_path.stem
        for marker in ("Prefeitura", "Câmara", "Camara"):
            if marker.lower() in stem.lower():
                return stem.split(marker, 1)[-1].strip(" -_")
        return None

    def _infer_tipo_orgao(self, workbook_path: Path) -> Optional[str]:
        name = workbook_path.stem.lower()
        if "prefeitura" in name:
            return "prefeitura"
        if "camara" in name or "câmara" in name:
            return "camara"
        return None

    def _infer_sat_numero(self, workbook_path: Path) -> Optional[str]:
        match = re.search(r"sat[_\s-]*(\d+)", workbook_path.stem, re.IGNORECASE)
        if match:
            return match.group(1)
        return None

    def _find_first_url_for_items(self, sheet, item_codes: list[str]) -> Optional[str]:
        wanted = {code.upper() for code in item_codes}
        for row_idx in range(self.config.metadata_row + 1, (sheet.max_row or 0) + 1):
            item_code = _clean_value(sheet[f"B{row_idx}"].value)
            if not item_code or item_code.upper() not in wanted:
                continue

            for column_idx in range(1, (sheet.max_column or 0) + 1):
                value = _clean_value(sheet.cell(row_idx, column_idx).value)
                if not value:
                    continue
                url = _extract_url(value)
                if url:
                    return url

        return None

    def _detect_available_sources(
        self,
        sheet,
        response_columns: list[tuple[str, str]],
    ) -> list[FonteType]:
        source_rows = {
            "site_orgao": ["1.1"],
            "portal_transparencia": ["1.2"],
            "esic": ["5.1", "5.2", "5.4", "5.5", "5.6", "5.7", "5.8"],
        }
        found: list[FonteType] = []
        for source_key, item_codes in source_rows.items():
            if self._has_available_item(sheet, response_columns, item_codes):
                found.append(source_key)
        return found

    def _has_available_item(
        self,
        sheet,
        response_columns: list[tuple[str, str]],
        item_codes: list[str],
    ) -> bool:
        wanted = {code.upper() for code in item_codes}
        for row_idx in range(self.config.metadata_row + 1, (sheet.max_row or 0) + 1):
            item_code = _clean_value(sheet[f"B{row_idx}"].value)
            if not item_code or item_code.upper() not in wanted:
                continue
            selected_status = self._select_reference_status(sheet, row_idx, response_columns)
            if selected_status is None:
                continue
            status = _normalize_status(selected_status[1])
            if status and status != "Nao se aplica":
                return True
        return False


def _normalize_text(value: object) -> str:
    cleaned = _clean_value(value)
    if not cleaned:
        return ""

    cleaned = unicodedata.normalize("NFKD", cleaned)
    cleaned = "".join(char for char in cleaned if not unicodedata.combining(char))
    cleaned = cleaned.lower()
    cleaned = cleaned.replace("-", "_").replace("/", "_")
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = re.sub(r"[^\w_]", "", cleaned)
    return cleaned.strip("_")


def _clean_value(value: object) -> Optional[str]:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None
    return text


def _as_number(value: object) -> Optional[float]:
    text = _clean_value(value)
    if text is None:
        return None
    normalized = text.replace("R$", "").replace(" ", "")
    if normalized.count(",") == 1 and normalized.count(".") == 0:
        normalized = normalized.replace(",", ".")
    elif normalized.count(",") >= 1 and normalized.count(".") >= 1:
        normalized = normalized.replace(".", "").replace(",", ".")
    else:
        normalized = normalized.replace(",", "")
    try:
        return float(normalized)
    except ValueError:
        return None


def _extract_url(value: str) -> Optional[str]:
    urls = _extract_urls(value)
    if not urls:
        return None
    return urls[0]


def _extract_urls(value: str) -> list[str]:
    matches = re.findall(r"https?://\S+", value, re.IGNORECASE)
    cleaned: list[str] = []
    seen: set[str] = set()
    for match in matches:
        url = match.rstrip(").,;")
        if url in seen:
            continue
        seen.add(url)
        cleaned.append(url)
    return cleaned


def _normalize_url_for_dedup(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    parsed = urlparse(raw)
    host = (parsed.netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    path = (parsed.path or "").rstrip("/")
    query = f"?{parsed.query}" if parsed.query else ""
    return f"{parsed.scheme.lower()}://{host}{path}{query}".rstrip("/")


def _entity_key(value: object) -> str:
    normalized = _normalize_text(value)
    return re.sub(r"[^a-z0-9]+", "", normalized)


def _is_crawlable_reference_url(value: str) -> bool:
    lowered = (value or "").strip().lower()
    if not lowered.startswith(("http://", "https://")):
        return False
    for extension in NON_CRAWLABLE_EXTENSIONS:
        if lowered.endswith(extension):
            return False
    return True


def _truncate_text(value: Optional[str], limit: int) -> Optional[str]:
    text = _clean_value(value)
    if not text:
        return None
    compact = " ".join(text.split())
    if len(compact) <= limit:
        return compact
    return compact[: limit - 3].rstrip() + "..."


def _normalize_status(value: Optional[str]) -> Optional[str]:
    normalized = _normalize_text(value)
    if not normalized:
        return None

    for raw, final in STATUS_MAP.items():
        if raw in normalized:
            return final
    return None


def _normalize_detail_status(value: Optional[str]) -> str:
    normalized = _normalize_status(value)
    if normalized:
        return normalized

    text = _clean_value(value)
    if not text:
        return "Nao informado"
    return text


def _normalize_fonte(value: Optional[str]) -> FonteType:
    normalized = _normalize_text(value)
    if not normalized:
        return "nao_informada"

    if "esic" in normalized or "e_sic" in normalized or "sic" == normalized:
        return "esic"
    if "portal" in normalized and "transparencia" in normalized:
        return "portal_transparencia"
    if "site" in normalized or "orgao" in normalized or "institucional" in normalized:
        return "site_orgao"
    return "nao_informada"


def list_parser_profiles() -> list[ParserProfileDefinition]:
    return list(PARSER_PROFILE_MAP.values())


def get_parser_profile_definition(profile: Optional[str]) -> ParserProfileDefinition:
    normalized_profile = (profile or "default").strip().lower() or "default"
    return PARSER_PROFILE_MAP.get(normalized_profile, PARSER_PROFILE_MAP["default"])


def resolve_parser_profile_for_workbook(
    workbook_path: Path,
    requested_profile: Optional[str] = None,
) -> str:
    normalized_profile = (requested_profile or "auto").strip().lower() or "auto"
    if normalized_profile not in AUTO_PROFILE_SELECTIONS:
        return get_parser_profile_definition(normalized_profile).key

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if _looks_like_financial_workbook(workbook):
            return "financial_dre"
        if _looks_like_checklist_workbook(workbook):
            return "extended"
    finally:
        workbook.close()
    return "extended"


def build_parser_config(
    profile: Optional[str] = None,
    allowed_groups_text: Optional[str] = None,
    allowed_status_text: Optional[str] = None,
    checklist_sheet_name: Optional[str] = None,
    metadata_row: Optional[int] = None,
) -> ParserConfig:
    normalized_profile = (profile or "default").strip().lower() or "default"
    profile_definition = PARSER_PROFILE_MAP.get(normalized_profile, PARSER_PROFILE_MAP["default"])
    if profile_definition.key == "auto":
        profile_definition = PARSER_PROFILE_MAP["extended"]

    if profile_definition.key == "financial_dre":
        normalized_sheet_selection = _normalize_sheet_selection_request(checklist_sheet_name)
        return ParserConfig(
            profile=profile_definition.key,
            allowed_groups=set(),
            allowed_status=set(),
            checklist_sheet_name=normalized_sheet_selection,
            metadata_row=max(1, int(metadata_row or 5)),
        )

    allowed_groups = _parse_csv_values(allowed_groups_text) or profile_definition.allowed_groups
    normalized_groups = sorted({group.strip() for group in allowed_groups if group.strip()})

    status_values = _parse_csv_values(allowed_status_text) or profile_definition.allowed_status
    normalized_statuses = _normalize_allowed_statuses(status_values) or profile_definition.allowed_status
    normalized_sheet_selection = _normalize_sheet_selection_request(checklist_sheet_name)

    return ParserConfig(
        profile=profile_definition.key,
        allowed_groups=set(normalized_groups or DEFAULT_ALLOWED_GROUPS),
        allowed_status=set(normalized_statuses or DEFAULT_ALLOWED_STATUS),
        checklist_sheet_name=normalized_sheet_selection,
        metadata_row=max(1, int(metadata_row or 5)),
    )


def _parse_csv_values(value: Optional[str]) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in value.split(",") if part.strip()]


def _parse_sheet_names(value: Optional[str]) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in re.split(r"[,;\n]+", value) if part.strip()]


def _normalize_sheet_selection_request(value: Optional[str]) -> str:
    sheet_names = _parse_sheet_names(value)
    if not sheet_names or _is_auto_sheet_selection(value):
        return "auto"
    return ", ".join(sheet_names)


def _looks_like_financial_workbook(workbook) -> bool:
    return any(_looks_like_financial_sheet_signature(sheet) for sheet in workbook.worksheets)


def _looks_like_checklist_workbook(workbook) -> bool:
    return any(_looks_like_checklist_sheet_signature(sheet, metadata_row=5) for sheet in workbook.worksheets)


def _looks_like_financial_sheet_signature(sheet) -> bool:
    normalized_title = _normalize_text(sheet.title)
    if normalized_title in FINANCIAL_MONTH_TITLES:
        return True
    if any(token in normalized_title for token in ("financeiro", "recebidos", "despesas", "construtivo", "flex")):
        if _looks_like_financial_data_signature(sheet):
            return True
    for row_idx in range(1, min((sheet.max_row or 0), 5) + 1):
        normalized_values = {
            _normalize_text(sheet.cell(row_idx, column_idx).value)
            for column_idx in range(1, min((sheet.max_column or 0), 12) + 1)
        }
        normalized_values.discard("")
        if {"cliente", "valor_r", "vencimento_fatura"}.issubset(normalized_values):
            return True
        if {"data", "produto"}.issubset(normalized_values) and any(
            token.startswith("debitos") for token in normalized_values
        ) and any(token.startswith("creditos") for token in normalized_values):
            return True
        row_text = " ".join(sorted(normalized_values))
        if "painel_financeiro" in row_text or "painel_de_controle" in row_text or "fluxo_financeiro" in row_text:
            return True
    return False


def _looks_like_financial_data_signature(sheet) -> bool:
    matched_rows = 0
    for row_idx in range(1, min((sheet.max_row or 0), 4) + 1):
        client = _clean_value(sheet.cell(row_idx, 5).value)
        amount = _clean_value(sheet.cell(row_idx, 7).value)
        if client and _as_number(amount) is not None:
            matched_rows += 1
    return matched_rows >= 2


def _looks_like_checklist_sheet_signature(sheet, metadata_row: int) -> bool:
    normalized_title = _normalize_text(sheet.title)
    if "checklist" in normalized_title:
        return True
    if (sheet.max_row or 0) < metadata_row:
        return False
    metadata_cells = next(
        sheet.iter_rows(min_row=metadata_row, max_row=metadata_row),
        [],
    )
    has_response_column = any(
        RESPONSE_HEADER_RE.search(_clean_value(cell.value) or "")
        for cell in metadata_cells
    )
    if not has_response_column:
        return False
    item_header = _normalize_text(sheet[f"B{metadata_row}"].value)
    description_header = _normalize_text(sheet[f"C{metadata_row}"].value)
    return item_header in {"item", "codigo", "item_codigo"} and description_header in {
        "descricao",
        "descricao_item",
        "descricao_do_item",
    }


def _is_auto_sheet_selection(value: Optional[str]) -> bool:
    normalized = _normalize_text(value)
    return normalized in AUTO_SHEET_SELECTIONS


def _normalize_allowed_statuses(values: list[str]) -> list[StatusType]:
    normalized: list[StatusType] = []
    for value in values:
        status = _normalize_status(value)
        if status and status not in normalized:
            normalized.append(status)
    return normalized


def _observation_expression_matches(expression: str, item_code: str) -> bool:
    expression_normalized = re.sub(r"\s+", "", expression.upper())
    item_normalized = re.sub(r"\s+", "", item_code.upper())
    if expression_normalized == item_normalized:
        return True

    expression_with_suffix = re.match(r"(\d+(?:\.\d+)?)([A-Z])$", expression_normalized)
    if expression_with_suffix and item_normalized == expression_with_suffix.group(1):
        return True

    tokens = re.findall(r"\d+(?:\.\d+)?[A-Z]?", expression_normalized)
    if item_normalized in tokens:
        return True

    range_match = re.match(r"(\d+(?:\.\d+)?)([A-Z])A([A-Z])", expression_normalized)
    item_match = re.match(r"(\d+(?:\.\d+)?)([A-Z])$", item_normalized)
    if range_match and item_match:
        prefix, start_letter, end_letter = range_match.groups()
        item_prefix, item_letter = item_match.groups()
        if prefix == item_prefix and start_letter <= item_letter <= end_letter:
            return True

    return False

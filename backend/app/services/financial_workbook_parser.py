from __future__ import annotations

import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse

from openpyxl import load_workbook

from ..models import (
    ChecklistParseResult,
    FinancialAnalysisResult,
    FinancialClientPeriodRollup,
    FinancialClientRollup,
    FinancialContractRollup,
    FinancialEntry,
    FinancialPeriodSummary,
    FinancialSectionSnapshot,
    FinancialStatementLine,
    ParserOptions,
    WorkbookContextLayer,
    WorkbookReferenceLink,
)
from .excel_parser import ParserConfig

MONTH_NAME_MAP = {
    "janeiro": (1, "Janeiro"),
    "fevereiro": (2, "Fevereiro"),
    "marco": (3, "Marco"),
    "março": (3, "Marco"),
    "abril": (4, "Abril"),
    "maio": (5, "Maio"),
    "junho": (6, "Junho"),
    "julho": (7, "Julho"),
    "agosto": (8, "Agosto"),
    "setembro": (9, "Setembro"),
    "outubro": (10, "Outubro"),
    "novembro": (11, "Novembro"),
    "dezembro": (12, "Dezembro"),
}

SECTION_TYPE_MAP = {
    "impostos": ("tax", "Impostos e encargos"),
    "despesas_com_pessoal": ("personnel", "Despesas com pessoal"),
    "custos_fixos": ("fixed_cost", "Custos fixos"),
    "custos_operacionais": ("operating_cost", "Custos operacionais"),
}

SUMMARY_FIELD_MAP = {
    "totalizacao_vbc_distribuidor": "vbc_total",
    "totalizacao_modulo_versa": "modulo_total",
    "despesas_globais": "global_expenses_total",
    "recebiveis_modulo_versa": "receivables_total",
    "resultado_final": "net_result",
    "saldo_dos_outros_meses": "carried_balance",
    "saldo_negativo_outros_meses": "carried_balance",
}

NON_CRAWLABLE_EXTENSIONS = {
    ".pdf",
    ".csv",
    ".xls",
    ".xlsx",
    ".ods",
    ".doc",
    ".docx",
    ".odt",
    ".zip",
    ".rar",
    ".7z",
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
}

AUTO_SHEET_SELECTIONS = {"", "*", "auto", "all", "todas", "todas_as_abas", "multi_aba"}
REALIZED_INVOICE_STATUSES = {"liquidado", "baixado"}
REALIZED_RECEIVABLE_STATUSES = REALIZED_INVOICE_STATUSES | {"pago", "recebido", "quitado"}


@dataclass
class RevenueBlockResult:
    receivable_entries: list[FinancialEntry]
    other_income_entries: list[FinancialEntry]
    subtotal_amount: Optional[float] = None
    total_amount: Optional[float] = None


@dataclass
class ScheduleParseResult:
    entries: list[FinancialEntry] = field(default_factory=list)
    realized_entries: list[FinancialEntry] = field(default_factory=list)
    pending_entries: list[FinancialEntry] = field(default_factory=list)
    reported_total: Optional[float] = None


@dataclass
class CashflowLedgerParseResult:
    expense_sections: list[FinancialSectionSnapshot] = field(default_factory=list)
    credit_entries: list[FinancialEntry] = field(default_factory=list)
    receivable_entries: list[FinancialEntry] = field(default_factory=list)
    debit_total: Optional[float] = None
    credit_total: Optional[float] = None
    receivable_total: Optional[float] = None


class FinancialWorkbookParser:
    def __init__(self, config: ParserConfig) -> None:
        self.config = config
        self._sheet_row_limits: dict[str, int] = {}

    def parse(self, workbook_path: Path, source_name: Optional[str] = None) -> ChecklistParseResult:
        result = ChecklistParseResult(
            grupos_permitidos=[],
            parser_options=ParserOptions(
                profile=self.config.profile,
                allowed_groups=[],
                allowed_status=[],
                checklist_sheet_name=self.config.checklist_sheet_name,
                metadata_row=self.config.metadata_row,
            ),
            tipo_orgao="financeiro",
        )
        workbook = load_workbook(workbook_path, data_only=True)
        try:
            sheets, missing_sheet_names = self._resolve_target_sheets(workbook)
            if not sheets:
                if missing_sheet_names:
                    result.warnings.append(
                        "Abas solicitadas nao encontradas: " + ", ".join(missing_sheet_names) + "."
                    )
                result.warnings.append("Nenhuma aba financeira compativel foi encontrada.")
                return result

            if missing_sheet_names:
                result.warnings.append(
                    "Abas solicitadas nao encontradas: " + ", ".join(missing_sheet_names) + "."
                )

            result.parser_options.checklist_sheet_names = [sheet.title for sheet in sheets]
            filename_hint = Path(source_name) if source_name else workbook_path
            result.orgao = self._infer_entity_name(filename_hint)
            result.reference_links = self._extract_reference_links(workbook, result.parser_options.checklist_sheet_names)

            ledger_bundle = self._resolve_ledger_sheet_bundle(workbook, sheets)
            if ledger_bundle and not any(self._looks_like_month_sheet(sheet) for sheet in sheets):
                return self._parse_ledger_workbook(
                    result=result,
                    sheets=ledger_bundle,
                    filename_hint=filename_hint,
                )

            result.periodo_analise = self._build_period_label(sheets, filename_hint)

            months: list[FinancialPeriodSummary] = []
            summary_notes: list[str] = []
            detected_entities: set[str] = set()
            for sheet in sheets:
                period = self._parse_month_sheet(sheet, filename_hint)
                if period is None:
                    result.warnings.append(
                        f"A aba '{sheet.title}' nao apresentou estrutura financeira suficiente para consolidacao."
                    )
                    continue
                months.append(period)
                summary_notes.extend(period.notes)
                for entry in self._iter_period_entries(period):
                    if entry.counterparty:
                        detected_entities.add(entry.counterparty)
                    if entry.owner_label:
                        detected_entities.add(entry.owner_label)

            if not months:
                result.warnings.append("Nao foi possivel extrair nenhum periodo financeiro consolidado.")
                return result

            if not result.orgao:
                result.orgao = self._guess_entity_from_periods(months, filename_hint)

            result.financial_analysis = self._build_financial_analysis(
                entity_name=result.orgao,
                months=months,
                summary_notes=summary_notes,
                detected_entities=detected_entities,
            )
            result.context_layers = self._build_context_layers(result.financial_analysis)
            result.warnings.extend(self._build_financial_warnings(result.financial_analysis))
            return result
        finally:
            workbook.close()

    def _resolve_target_sheets(self, workbook) -> tuple[list, list[str]]:
        if _is_auto_sheet_selection(self.config.checklist_sheet_name):
            month_sheets = [sheet for sheet in workbook.worksheets if self._looks_like_month_sheet(sheet)]
            if month_sheets:
                return month_sheets, []
            ledger_bundle = self._resolve_ledger_sheet_bundle(workbook)
            if ledger_bundle:
                return ledger_bundle, []
            return [], []

        requested_names = _parse_sheet_names(self.config.checklist_sheet_name)
        selected = []
        selected_normalized: set[str] = set()
        missing: list[str] = []
        for requested_name in requested_names:
            sheet = self._find_sheet_by_name(workbook, requested_name)
            if sheet is None:
                missing.append(requested_name)
                continue
            normalized_title = _normalize_text(sheet.title)
            if normalized_title in selected_normalized:
                continue
            selected.append(sheet)
            selected_normalized.add(normalized_title)
        return selected, missing

    def _find_sheet_by_name(self, workbook, requested_name: str):
        if requested_name in workbook.sheetnames:
            return workbook[requested_name]
        requested_normalized = _normalize_text(requested_name)
        for sheet_name in workbook.sheetnames:
            if _normalize_text(sheet_name) == requested_normalized:
                return workbook[sheet_name]
        return None

    def _looks_like_financial_sheet(self, sheet) -> bool:
        return (
            self._looks_like_month_sheet(sheet)
            or self._looks_like_financial_schedule_sheet(sheet)
            or self._looks_like_cashflow_ledger_sheet(sheet)
            or self._looks_like_inventory_sheet(sheet)
        )

    def _looks_like_month_sheet(self, sheet) -> bool:
        normalized_title = _normalize_text(sheet.title)
        if normalized_title in MONTH_NAME_MAP:
            return True
        for row_idx in range(1, min((sheet.max_row or 0), 12) + 1):
            row_text = " ".join(
                value
                for value in (_clean_value(sheet.cell(row_idx, column_idx).value) for column_idx in range(1, 10))
                if value
            ).lower()
            if "painel de controle" in row_text:
                return True
        return False

    def _looks_like_financial_schedule_sheet(self, sheet) -> bool:
        normalized_title = _normalize_text(sheet.title)
        for row_idx in range(1, min((sheet.max_row or 0), 5) + 1):
            normalized_values = {
                column_idx: _normalize_text(sheet.cell(row_idx, column_idx).value)
                for column_idx in range(1, min((sheet.max_column or 0), 12) + 1)
            }
            normalized_set = {value for value in normalized_values.values() if value}
            if {"cliente", "valor_r", "vencimento_fatura"}.issubset(normalized_set):
                return True
        if "recebido" in normalized_title or "financeiro" in normalized_title:
            data_rows = sum(1 for row_idx in range(1, 5) if self._looks_like_schedule_data_row(sheet, row_idx))
            if data_rows >= 2:
                return True
        return False

    def _looks_like_cashflow_ledger_sheet(self, sheet) -> bool:
        normalized_title = _normalize_text(sheet.title)
        for row_idx in range(1, min((sheet.max_row or 0), 4) + 1):
            normalized_values = {
                column_idx: _normalize_text(sheet.cell(row_idx, column_idx).value)
                for column_idx in range(1, min((sheet.max_column or 0), 6) + 1)
            }
            normalized_set = {value for value in normalized_values.values() if value}
            if {"data", "produto"}.issubset(normalized_set) and any(
                token.startswith("debitos") for token in normalized_set
            ) and any(token.startswith("creditos") for token in normalized_set):
                return True
        return "despesas" in normalized_title and "modulo_versa" in normalized_title

    def _looks_like_inventory_sheet(self, sheet) -> bool:
        normalized_title = _normalize_text(sheet.title)
        if "construtivo" not in normalized_title and "flex" not in normalized_title:
            return False
        header_values = {
            _normalize_text(sheet.cell(2, column_idx).value)
            for column_idx in range(1, min((sheet.max_column or 0), 9) + 1)
        }
        return {"n_lote", "codigo_sistema", "situacao"}.issubset(header_values)

    def _looks_like_schedule_data_row(self, sheet, row_idx: int) -> bool:
        lot = _clean_value(sheet.cell(row_idx, 1).value)
        client = _clean_value(sheet.cell(row_idx, 5).value)
        amount = _as_number(sheet.cell(row_idx, 7).value)
        contract_date = _stringify_date(sheet.cell(row_idx, 4).value)
        due_date = _stringify_date(sheet.cell(row_idx, 8).value)
        if not client or amount is None:
            return False
        if not (contract_date or due_date):
            return False
        return bool(lot or _clean_value(sheet.cell(row_idx, 3).value))

    def _resolve_ledger_sheet_bundle(self, workbook, candidate_sheets: Optional[list] = None) -> list:
        sheets = list(candidate_sheets) if candidate_sheets is not None else list(workbook.worksheets)
        schedule_sheets: list = []
        received_sheets: list = []
        ledger_sheets: list = []
        inventory_sheets: list = []

        for sheet in sheets:
            normalized_title = _normalize_text(sheet.title)
            if self._looks_like_financial_schedule_sheet(sheet):
                if "recebido" in normalized_title:
                    received_sheets.append(sheet)
                else:
                    schedule_sheets.append(sheet)
                continue
            if self._looks_like_cashflow_ledger_sheet(sheet):
                ledger_sheets.append(sheet)
                continue
            if self._looks_like_inventory_sheet(sheet):
                inventory_sheets.append(sheet)

        if not schedule_sheets and not received_sheets and not ledger_sheets:
            return []

        ordered: list = []
        seen_titles: set[str] = set()
        for bucket in (schedule_sheets, received_sheets, ledger_sheets, inventory_sheets):
            for sheet in bucket:
                normalized_title = _normalize_text(sheet.title)
                if normalized_title in seen_titles:
                    continue
                ordered.append(sheet)
                seen_titles.add(normalized_title)
        return ordered

    def _parse_ledger_workbook(
        self,
        result: ChecklistParseResult,
        sheets: list,
        filename_hint: Path,
    ) -> ChecklistParseResult:
        schedule_sheet = next(
            (
                sheet
                for sheet in sheets
                if self._looks_like_financial_schedule_sheet(sheet)
                and "recebido" not in _normalize_text(sheet.title)
            ),
            None,
        )
        received_sheet = next(
            (
                sheet
                for sheet in sheets
                if self._looks_like_financial_schedule_sheet(sheet)
                and "recebido" in _normalize_text(sheet.title)
            ),
            None,
        )
        ledger_sheet = next((sheet for sheet in sheets if self._looks_like_cashflow_ledger_sheet(sheet)), None)
        inventory_sheets = [sheet for sheet in sheets if self._looks_like_inventory_sheet(sheet)]

        ledger_entity_name = self._guess_ledger_entity_name(filename_hint, inventory_sheets)
        if ledger_entity_name:
            result.orgao = ledger_entity_name

        schedule_data = (
            self._parse_schedule_entries(schedule_sheet)
            if schedule_sheet is not None
            else ScheduleParseResult()
        )
        received_data = (
            self._parse_schedule_entries(received_sheet, assume_realized=True)
            if received_sheet is not None
            else ScheduleParseResult()
        )
        ledger_data = (
            self._parse_cashflow_ledger_entries(ledger_sheet)
            if ledger_sheet is not None
            else CashflowLedgerParseResult()
        )

        realized_entries = (
            list(received_data.realized_entries or received_data.entries)
            or list(schedule_data.realized_entries)
        )
        unmatched_credit_entries = self._select_unmatched_credit_entries(
            reference_entries=list(realized_entries) + list(schedule_data.realized_entries),
            credit_entries=ledger_data.credit_entries,
        )
        open_entries = list(schedule_data.pending_entries)
        if not open_entries and ledger_data.receivable_entries:
            open_entries = list(ledger_data.receivable_entries)

        sections: list[FinancialSectionSnapshot] = []
        if realized_entries:
            sections.append(
                FinancialSectionSnapshot(
                    section_key="receivables",
                    title="Recebimentos realizados",
                    total_amount=_sum_amount(entry.amount for entry in realized_entries),
                    entry_count=len(realized_entries),
                    entries=realized_entries,
                )
            )
        if open_entries:
            sections.append(
                FinancialSectionSnapshot(
                    section_key="receivable_open",
                    title="Carteira em aberto",
                    total_amount=_sum_amount(entry.amount for entry in open_entries),
                    entry_count=len(open_entries),
                    entries=open_entries,
                )
            )
        if unmatched_credit_entries:
            sections.append(
                FinancialSectionSnapshot(
                    section_key="other_income",
                    title="Entradas complementares em fluxo",
                    total_amount=_sum_amount(entry.amount for entry in unmatched_credit_entries),
                    entry_count=len(unmatched_credit_entries),
                    entries=unmatched_credit_entries,
                )
            )
        sections.extend(ledger_data.expense_sections)

        section_totals = self._aggregate_section_totals(ledger_data.expense_sections)
        realized_total = _sum_amount(entry.amount for entry in realized_entries)
        open_total = _sum_amount(entry.amount for entry in open_entries)
        other_income_total = _sum_amount(entry.amount for entry in unmatched_credit_entries)
        gross_total = (
            schedule_data.reported_total
            or _sum_amount(entry.amount for entry in schedule_data.entries)
            or _sum_amount(
                amount
                for amount in (
                    realized_total,
                    open_total,
                    other_income_total,
                )
            )
        )
        global_expenses_total = _sum_amount(
            (
                section_totals.get("tax"),
                section_totals.get("personnel"),
                section_totals.get("fixed_cost"),
                section_totals.get("operating_cost"),
            )
        )
        net_result = None
        if global_expenses_total is not None:
            net_result = _first_number(realized_total, 0.0) + _first_number(other_income_total, 0.0) - global_expenses_total

        notes = self._build_ledger_notes(
            schedule_sheet=schedule_sheet,
            received_sheet=received_sheet,
            ledger_sheet=ledger_sheet,
            schedule_data=schedule_data,
            received_data=received_data,
            ledger_data=ledger_data,
            unmatched_credit_entries=unmatched_credit_entries,
            open_total=open_total,
        )
        period = FinancialPeriodSummary(
            sheet_name="Consolidado",
            period_label=self._build_ledger_period_label(
                schedule_data=schedule_data,
                received_data=received_data,
                ledger_data=ledger_data,
                filename_hint=filename_hint,
            ),
            year=self._resolve_ledger_year(
                schedule_data=schedule_data,
                received_data=received_data,
                ledger_data=ledger_data,
                filename_hint=filename_hint,
            ),
            gross_revenue_total=gross_total,
            receivables_total=realized_total,
            other_income_total=other_income_total,
            taxes_total=section_totals.get("tax"),
            personnel_total=section_totals.get("personnel"),
            fixed_costs_total=section_totals.get("fixed_cost"),
            operating_costs_total=section_totals.get("operating_cost"),
            global_expenses_total=global_expenses_total,
            net_result=net_result,
            pending_entry_count=sum(1 for entry in open_entries if entry.amount is not None),
            sections=sections,
            notes=notes,
        )

        detected_entities: set[str] = set()
        for entry in self._iter_period_entries(period):
            if entry.counterparty:
                detected_entities.add(entry.counterparty)
            if entry.owner_label:
                detected_entities.add(entry.owner_label)

        inventory_layers, inventory_entities = self._build_inventory_layers(inventory_sheets)
        detected_entities.update(inventory_entities)

        if result.orgao is None:
            result.orgao = self._guess_entity_from_periods([period], filename_hint)

        result.periodo_analise = period.period_label
        result.financial_analysis = self._build_financial_analysis(
            entity_name=result.orgao,
            months=[period],
            summary_notes=notes,
            detected_entities=detected_entities,
        )
        result.context_layers = self._build_context_layers(result.financial_analysis) + inventory_layers
        result.warnings.extend(self._build_financial_warnings(result.financial_analysis))
        return result

    def _parse_schedule_entries(self, sheet, assume_realized: bool = False) -> ScheduleParseResult:
        result = ScheduleParseResult()
        if sheet is None:
            return result

        header_row = self._find_schedule_header_row(sheet)
        data_start_row = header_row + 1 if header_row is not None else 1
        for row_idx in range(data_start_row, self._row_limit(sheet) + 1):
            lot_label = _clean_value(sheet.cell(row_idx, 1).value)
            client_name = _clean_value(sheet.cell(row_idx, 5).value)
            amount = _as_number(sheet.cell(row_idx, 7).value)
            invoice_status = _clean_value(sheet.cell(row_idx, 9).value)
            payment_date = _stringify_date(sheet.cell(row_idx, 10).value)
            due_date = _stringify_date(sheet.cell(row_idx, 8).value)
            contract_status = _clean_value(sheet.cell(row_idx, 3).value)
            contract_date = _stringify_date(sheet.cell(row_idx, 4).value)
            term = _clean_value(sheet.cell(row_idx, 6).value)
            observation = _clean_value(sheet.cell(row_idx, 11).value)

            if _normalize_text(lot_label) in {"total", "total_"}:
                if amount is not None:
                    result.reported_total = amount
                continue
            if not any((lot_label, client_name, amount is not None, invoice_status, observation)):
                continue
            if client_name is None or amount is None:
                continue

            entry = FinancialEntry(
                entry_type="receivable",
                sheet_name=sheet.title,
                description=self._schedule_entry_description(client_name, observation, lot_label),
                amount=amount,
                status="Liquidado" if assume_realized else (invoice_status or contract_status),
                date=payment_date or contract_date,
                due_date=due_date,
                counterparty=client_name,
                unit=lot_label,
                notes=self._join_notes(contract_status, term, observation),
                contract_label=self._build_contract_label(
                    client_name=client_name,
                    unit=lot_label,
                    contract_start_date=contract_date,
                    contract_end_date=None,
                    term=term,
                ),
                contract_start_date=contract_date,
            )
            result.entries.append(entry)

            normalized_status = _normalize_text(invoice_status or "")
            if assume_realized or normalized_status in REALIZED_INVOICE_STATUSES:
                result.realized_entries.append(entry)
            else:
                pending_status = invoice_status or "A receber"
                result.pending_entries.append(entry.copy(update={"status": pending_status}))

        if result.reported_total is None:
            result.reported_total = _sum_amount(entry.amount for entry in result.entries)
        return result

    def _find_schedule_header_row(self, sheet) -> Optional[int]:
        for row_idx in range(1, min((sheet.max_row or 0), 5) + 1):
            normalized_values = {
                _normalize_text(sheet.cell(row_idx, column_idx).value)
                for column_idx in range(1, min((sheet.max_column or 0), 12) + 1)
            }
            if {"cliente", "valor_r", "vencimento_fatura"}.issubset(normalized_values):
                return row_idx
        return None

    def _schedule_entry_description(
        self,
        client_name: str,
        observation: Optional[str],
        lot_label: Optional[str],
    ) -> str:
        if observation:
            return f"{client_name} | {observation}"
        if lot_label:
            return f"{client_name} | {lot_label}"
        return client_name

    def _parse_cashflow_ledger_entries(self, sheet) -> CashflowLedgerParseResult:
        result = CashflowLedgerParseResult()
        if sheet is None:
            return result

        sections_by_key: dict[str, FinancialSectionSnapshot] = {}
        for row_idx in range(3, self._row_limit(sheet) + 1):
            entry_date = _stringify_date(sheet.cell(row_idx, 1).value)
            description = _clean_value(sheet.cell(row_idx, 2).value)
            debit_amount = _as_number(sheet.cell(row_idx, 3).value)
            credit_amount = _as_number(sheet.cell(row_idx, 4).value)
            receivable_amount = _as_number(sheet.cell(row_idx, 5).value)

            if _normalize_text(description) == "total":
                result.debit_total = debit_amount
                result.credit_total = credit_amount
                result.receivable_total = receivable_amount
                continue
            if not any((description, debit_amount is not None, credit_amount is not None, receivable_amount is not None)):
                continue
            if not description:
                continue

            if debit_amount is not None:
                section_key = self._classify_ledger_expense_type(description)
                section_title = dict(SECTION_TYPE_MAP.values()).get(section_key)
                if section_title is None:
                    section_title = "Custos operacionais"
                snapshot = sections_by_key.setdefault(
                    section_key,
                    FinancialSectionSnapshot(
                        section_key=section_key,
                        title=section_title,
                    ),
                )
                snapshot.entries.append(
                    FinancialEntry(
                        entry_type=section_key,
                        sheet_name=sheet.title,
                        description=description,
                        amount=debit_amount,
                        status="Pago",
                        date=entry_date,
                        due_date=entry_date,
                        counterparty=description,
                    )
                )
                snapshot.entry_count += 1

            if credit_amount is not None:
                result.credit_entries.append(
                    FinancialEntry(
                        entry_type="other_income",
                        sheet_name=sheet.title,
                        description=description,
                        amount=credit_amount,
                        status="Liquidado",
                        date=entry_date,
                        counterparty=description,
                    )
                )

            if receivable_amount is not None:
                result.receivable_entries.append(
                    FinancialEntry(
                        entry_type="receivable",
                        sheet_name=sheet.title,
                        description=description,
                        amount=receivable_amount,
                        status="A receber",
                        date=entry_date,
                        counterparty=description,
                    )
                )

        for snapshot in sections_by_key.values():
            snapshot.total_amount = _sum_amount(entry.amount for entry in snapshot.entries)
        result.expense_sections = list(sections_by_key.values())
        if result.debit_total is None:
            result.debit_total = _sum_amount(
                snapshot.total_amount for snapshot in result.expense_sections if snapshot.total_amount is not None
            )
        if result.credit_total is None:
            result.credit_total = _sum_amount(entry.amount for entry in result.credit_entries)
        if result.receivable_total is None:
            result.receivable_total = _sum_amount(entry.amount for entry in result.receivable_entries)
        return result

    def _classify_ledger_expense_type(self, description: str) -> str:
        normalized = _normalize_text(description)
        tax_tokens = (
            "imposto",
            "tribut",
            "simples",
            "iss",
            "icms",
            "inss",
            "fgts",
            "darf",
            "taxa",
            "municipal",
            "federal",
        )
        personnel_tokens = (
            "salario",
            "pro_labore",
            "prolabore",
            "secretaria",
            "funcionario",
            "funcionaria",
            "remessa",
            "pagamento_",
        )
        fixed_tokens = (
            "vivo",
            "edp",
            "brk",
            "aluguel",
            "agua",
            "energia",
            "hospedagem",
            "dominio",
            "internet",
            "telefone",
            "consultoria",
            "nic_br",
        )
        if any(token in normalized for token in tax_tokens):
            return "tax"
        if any(token in normalized for token in personnel_tokens):
            return "personnel"
        if any(token in normalized for token in fixed_tokens):
            return "fixed_cost"
        return "operating_cost"

    def _select_unmatched_credit_entries(
        self,
        reference_entries: list[FinancialEntry],
        credit_entries: list[FinancialEntry],
    ) -> list[FinancialEntry]:
        if not credit_entries:
            return []
        remaining = Counter(
            (
                self._normalized_entry_date(entry),
                round(entry.amount or 0.0, 2),
            )
            for entry in reference_entries
            if entry.amount is not None
        )
        unmatched: list[FinancialEntry] = []
        for entry in credit_entries:
            if entry.amount is None:
                continue
            signature = (
                self._normalized_entry_date(entry),
                round(entry.amount, 2),
            )
            if remaining.get(signature, 0) > 0:
                remaining[signature] -= 1
                continue
            unmatched.append(entry)
        return unmatched

    def _normalized_entry_date(self, entry: FinancialEntry) -> str:
        return entry.date or entry.due_date or ""

    def _build_ledger_notes(
        self,
        schedule_sheet,
        received_sheet,
        ledger_sheet,
        schedule_data: ScheduleParseResult,
        received_data: ScheduleParseResult,
        ledger_data: CashflowLedgerParseResult,
        unmatched_credit_entries: list[FinancialEntry],
        open_total: Optional[float],
    ) -> list[str]:
        notes: list[str] = []
        if schedule_sheet is not None:
            notes.append(
                f"{schedule_sheet.title}: painel com {len(schedule_data.entries)} lancamento(s) e carteira prevista de {_format_currency(schedule_data.reported_total)}."
            )
        if received_sheet is not None:
            notes.append(
                f"{received_sheet.title}: {len(received_data.entries)} recebimento(s) realizados, somando {_format_currency(received_data.reported_total)}."
            )
        if ledger_sheet is not None:
            notes.append(
                f"{ledger_sheet.title}: fluxo com {_format_currency(ledger_data.debit_total)} em pagamentos e {_format_currency(ledger_data.credit_total)} em creditos."
            )
        if (
            received_sheet is not None
            and schedule_data.realized_entries
            and received_data.reported_total is not None
            and abs(received_data.reported_total - _first_number(_sum_amount(entry.amount for entry in schedule_data.realized_entries), 0.0)) > 0.01
        ):
            notes.append(
                "Os valores realizados marcados no painel financeiro nao coincidem integralmente com a aba de recebidos; o parser priorizou a aba RECEBIDOS para receita realizada."
            )
        if unmatched_credit_entries:
            notes.append(
                f"Foram isoladas {len(unmatched_credit_entries)} entrada(s) complementar(es) no fluxo financeiro que nao aparecem na trilha principal de recebimentos."
            )
        if open_total is not None:
            notes.append(f"Carteira em aberto identificada em {_format_currency(open_total)}.")
        if ledger_data.receivable_total is not None and (open_total is None or abs(ledger_data.receivable_total - open_total) > 0.01):
            notes.append(
                f"O fluxo contabil registra {_format_currency(ledger_data.receivable_total)} em recebiveis complementares."
            )
        return notes

    def _build_ledger_period_label(
        self,
        schedule_data: ScheduleParseResult,
        received_data: ScheduleParseResult,
        ledger_data: CashflowLedgerParseResult,
        filename_hint: Path,
    ) -> str:
        dates = self._collect_date_values(
            list(schedule_data.entries)
            + list(received_data.entries)
            + list(ledger_data.credit_entries)
            + list(ledger_data.receivable_entries)
        )
        if not dates:
            fallback_year = self._detect_year_from_path(filename_hint)
            return str(fallback_year) if fallback_year else "Consolidado financeiro"
        start_date = min(dates)
        end_date = max(dates)
        if start_date.year == end_date.year and start_date.month == end_date.month:
            return start_date.strftime("%m/%Y")
        return f"{start_date.strftime('%m/%Y')} a {end_date.strftime('%m/%Y')}"

    def _resolve_ledger_year(
        self,
        schedule_data: ScheduleParseResult,
        received_data: ScheduleParseResult,
        ledger_data: CashflowLedgerParseResult,
        filename_hint: Path,
    ) -> Optional[int]:
        dates = self._collect_date_values(
            list(schedule_data.entries)
            + list(received_data.entries)
            + list(ledger_data.credit_entries)
            + list(ledger_data.receivable_entries)
        )
        if dates:
            return max(dates).year
        return self._detect_year_from_path(filename_hint)

    def _collect_date_values(self, entries: list[FinancialEntry]) -> list[datetime]:
        dates: list[datetime] = []
        for entry in entries:
            for candidate in (entry.date, entry.due_date):
                if not candidate:
                    continue
                try:
                    dates.append(datetime.fromisoformat(candidate))
                    break
                except ValueError:
                    continue
        return dates

    def _build_inventory_layers(self, sheets: list) -> tuple[list[WorkbookContextLayer], set[str]]:
        layers: list[WorkbookContextLayer] = []
        detected_entities: set[str] = set()
        for sheet in sheets:
            status_counter: Counter[str] = Counter()
            holder_counter: Counter[str] = Counter()
            row_count = 0
            for row_idx in range(3, self._row_limit(sheet) + 1):
                lot_label = _clean_value(sheet.cell(row_idx, 1).value)
                status = _clean_value(sheet.cell(row_idx, 5).value)
                holder = _clean_value(sheet.cell(row_idx, 7).value)
                if not lot_label and not status and not holder:
                    continue
                if not lot_label:
                    continue
                row_count += 1
                status_counter[status or "Nao informado"] += 1
                if holder:
                    holder_counter[holder] += 1
                    detected_entities.add(holder)
            if not row_count:
                continue
            top_status = ", ".join(
                f"{label}: {count}"
                for label, count in status_counter.most_common(4)
            )
            layers.append(
                WorkbookContextLayer(
                    layer_type="registry_snapshot",
                    sheet_name=sheet.title,
                    title=f"Carteira patrimonial de {sheet.title}",
                    summary=(
                        f"{row_count} registro(s) identificados, com destaque para {top_status or 'status nao informado'}."
                    ),
                    details=[
                        f"Status {label}: {count}"
                        for label, count in status_counter.most_common(6)
                    ]
                    + [
                        f"Contraparte recorrente: {label} ({count})"
                        for label, count in holder_counter.most_common(4)
                    ],
                )
            )
        return layers, detected_entities

    def _guess_ledger_entity_name(self, filename_hint: Path, inventory_sheets: list) -> Optional[str]:
        for sheet in inventory_sheets:
            title_cell = _clean_value(sheet.cell(1, 1).value)
            if title_cell:
                return _truncate_text(title_cell.title(), 120)
        return self._infer_entity_name(filename_hint)

    def _parse_month_sheet(self, sheet, filename_hint: Path) -> Optional[FinancialPeriodSummary]:
        revenue_block = self._parse_revenue_block(sheet)
        permuta_snapshot = self._parse_permuta_snapshot(sheet)
        debt_snapshot = self._parse_debt_snapshot(sheet)
        section_snapshots = self._parse_financial_sections(sheet)
        summary_values = self._extract_summary_values(sheet)

        sections = []
        if revenue_block.receivable_entries:
            sections.append(
                FinancialSectionSnapshot(
                    section_key="receivables",
                    title="Recebiveis operacionais",
                    total_amount=revenue_block.total_amount
                    or _sum_amount(entry.amount for entry in revenue_block.receivable_entries),
                    entry_count=len(revenue_block.receivable_entries),
                    entries=revenue_block.receivable_entries,
                )
            )
        if revenue_block.other_income_entries:
            sections.append(
                FinancialSectionSnapshot(
                    section_key="other_income",
                    title="Entradas complementares",
                    total_amount=_sum_amount(entry.amount for entry in revenue_block.other_income_entries),
                    entry_count=len(revenue_block.other_income_entries),
                    entries=revenue_block.other_income_entries,
                )
            )
        if permuta_snapshot:
            sections.append(permuta_snapshot)
        if debt_snapshot:
            sections.append(debt_snapshot)
        sections.extend(section_snapshots)

        if not sections and not summary_values:
            return None

        section_totals = self._aggregate_section_totals(section_snapshots)
        month_year = self._detect_sheet_year(sheet, filename_hint)
        month_label = self._month_label(sheet.title)
        receivables_total = summary_values.get("receivables_total") or revenue_block.total_amount
        gross_revenue_total = revenue_block.subtotal_amount or receivables_total
        other_income_total = _sum_amount(entry.amount for entry in revenue_block.other_income_entries)
        global_expenses_total = (
            summary_values.get("global_expenses_total")
            or _first_number(summary_values.get("vbc_total"), 0.0) + _first_number(summary_values.get("modulo_total"), 0.0)
            or _sum_amount(
                (
                    section_totals.get("tax"),
                    section_totals.get("personnel"),
                    section_totals.get("fixed_cost"),
                    section_totals.get("operating_cost"),
                )
            )
        )
        net_result = summary_values.get("net_result")
        if net_result is None and receivables_total is not None and global_expenses_total is not None:
            net_result = receivables_total + global_expenses_total

        pending_entry_count = sum(
            1
            for entry in self._iter_section_entries(sections)
            if _is_pending_status(entry.status)
        )
        notes = self._build_period_notes(
            sheet_name=sheet.title,
            revenue_block=revenue_block,
            summary_values=summary_values,
            section_totals=section_totals,
        )

        return FinancialPeriodSummary(
            sheet_name=sheet.title,
            period_label=f"{month_label}/{month_year}" if month_year else month_label,
            year=month_year,
            gross_revenue_total=gross_revenue_total,
            receivables_total=receivables_total,
            other_income_total=other_income_total,
            permuta_balance=permuta_snapshot.total_amount if permuta_snapshot else None,
            debt_outstanding=self._resolve_debt_outstanding(debt_snapshot),
            taxes_total=section_totals.get("tax"),
            personnel_total=section_totals.get("personnel"),
            fixed_costs_total=section_totals.get("fixed_cost"),
            operating_costs_total=section_totals.get("operating_cost"),
            vbc_total=summary_values.get("vbc_total"),
            modulo_total=summary_values.get("modulo_total"),
            global_expenses_total=global_expenses_total,
            net_result=net_result,
            carried_balance=summary_values.get("carried_balance"),
            closing_total=summary_values.get("closing_total"),
            pending_entry_count=pending_entry_count,
            sections=sections,
            notes=notes,
        )

    def _parse_revenue_block(self, sheet) -> RevenueBlockResult:
        header_row = None
        column_map: dict[str, int] = {}
        for row_idx in range(1, min((sheet.max_row or 0), 12) + 1):
            row_map = self._map_header_columns(sheet, row_idx)
            if "client" in row_map and "amount" in row_map and "payment_status" in row_map:
                header_row = row_idx
                column_map = row_map
                break

        receivable_entries: list[FinancialEntry] = []
        other_income_entries: list[FinancialEntry] = []
        subtotal_amount: Optional[float] = None
        total_amount: Optional[float] = None

        if header_row is not None:
            for row_idx in range(header_row + 1, min((sheet.max_row or 0), 30) + 1):
                row_values = self._row_values(sheet, row_idx, max_columns=max(column_map.values()) + 2)
                row_text = " ".join(row_values.values()).lower()
                if not row_values:
                    continue
                if "cliente" in row_text and "permuta" in row_text and "saldo" in row_text:
                    break
                if "divida" in row_text and "total" in row_text:
                    break
                if "vbc distribuidor" in row_text or "modulo versa" in row_text:
                    break

                numeric_total = _as_number(sheet.cell(row_idx, column_map["amount"]).value)
                leading_label = next(
                    (
                        value
                        for column_idx, value in sorted(row_values.items(), key=lambda item: int(item[0]))
                        if int(column_idx) < column_map["client"]
                    ),
                    None,
                )
                trailing_label = _clean_value(sheet.cell(row_idx, column_map["client"]).value)
                label_text = leading_label or trailing_label
                normalized_label = _normalize_text(label_text)
                if normalized_label == "subtotal" and numeric_total is not None:
                    subtotal_amount = numeric_total
                    continue
                if normalized_label == "total" and numeric_total is not None:
                    total_amount = numeric_total
                    continue

                client_name = _clean_value(sheet.cell(row_idx, column_map["client"]).value)
                amount = _as_number(sheet.cell(row_idx, column_map["amount"]).value)
                if not client_name or amount is None:
                    continue

                contract_status = self._cell_text(sheet, row_idx, column_map.get("contract_status"))
                payment_status = self._cell_text(sheet, row_idx, column_map.get("payment_status"))
                contract_start_date = self._cell_date(sheet, row_idx, column_map.get("start_date"))
                contract_end_date = self._cell_date(sheet, row_idx, column_map.get("due_date")) or self._cell_text(
                    sheet,
                    row_idx,
                    column_map.get("due_date"),
                )
                unit_label = self._cell_text(sheet, row_idx, column_map.get("unit"))
                contract_term = self._cell_text(sheet, row_idx, column_map.get("term"))
                notes = self._join_notes(
                    unit_label,
                    contract_status,
                    contract_term,
                    self._cell_text(sheet, row_idx, column_map.get("extra_status")),
                )
                is_permuta = "permuta" in _normalize_text(contract_status) or "permuta" in _normalize_text(payment_status)
                if is_permuta:
                    continue
                receivable_entries.append(
                    FinancialEntry(
                        entry_type="receivable",
                        sheet_name=sheet.title,
                        description=client_name,
                        amount=amount,
                        status=payment_status or contract_status,
                        date=contract_start_date,
                        due_date=contract_end_date,
                        counterparty=client_name,
                        unit=unit_label,
                        notes=notes,
                        contract_label=self._build_contract_label(
                            client_name=client_name,
                            unit=unit_label,
                            contract_start_date=contract_start_date,
                            contract_end_date=contract_end_date,
                            term=contract_term,
                        ),
                        contract_start_date=contract_start_date,
                        contract_end_date=contract_end_date,
                    )
                )

        aux_header = self._find_auxiliary_revenue_header(sheet)
        if aux_header:
            desc_col, date_col, amount_col, status_col, header_row = aux_header
            for row_idx in range(header_row + 1, min((sheet.max_row or 0), 30) + 1):
                row_values = self._row_values(sheet, row_idx, max_columns=max(amount_col, status_col))
                row_text = " ".join(row_values.values()).lower()
                if not row_values:
                    continue
                if "cliente" in row_text and "permuta" in row_text and "saldo" in row_text:
                    break
                if "divida" in row_text and "total" in row_text:
                    break
                if "vbc distribuidor" in row_text or "modulo versa" in row_text:
                    break

                label = _clean_value(sheet.cell(row_idx, desc_col).value)
                amount = _as_number(sheet.cell(row_idx, amount_col).value)
                normalized_label = _normalize_text(label)
                if normalized_label == "subtotal" and amount is not None:
                    subtotal_amount = subtotal_amount or amount
                    continue
                if normalized_label == "total" and amount is not None:
                    total_amount = total_amount or amount
                    continue
                if not label:
                    continue
                if amount is None and not _clean_value(sheet.cell(row_idx, status_col).value):
                    continue
                other_income_entries.append(
                    FinancialEntry(
                        entry_type="other_income",
                        sheet_name=sheet.title,
                        description=label,
                        amount=amount,
                        status=_clean_value(sheet.cell(row_idx, status_col).value),
                        date=_stringify_date(sheet.cell(row_idx, date_col).value),
                        counterparty=label,
                    )
                )

        return RevenueBlockResult(
            receivable_entries=receivable_entries,
            other_income_entries=other_income_entries,
            subtotal_amount=subtotal_amount,
            total_amount=total_amount,
        )

    def _map_header_columns(self, sheet, row_idx: int) -> dict[str, int]:
        mapping: dict[str, int] = {}
        status_columns: list[int] = []
        for column_idx in range(1, min((sheet.max_column or 0), 12) + 1):
            normalized = _normalize_text(sheet.cell(row_idx, column_idx).value)
            if not normalized:
                continue
            if normalized == "unidade":
                mapping["unit"] = column_idx
            elif normalized == "cliente":
                mapping["client"] = column_idx
            elif normalized in {"data", "locacao", "locacao_data"} or "locacao" in normalized:
                mapping["start_date"] = column_idx
            elif "fim_do_contrato" in normalized or "vencimento" in normalized:
                mapping["due_date"] = column_idx
            elif "prazo" in normalized:
                mapping["term"] = column_idx
            elif "valor" in normalized:
                mapping["amount"] = column_idx
            elif normalized == "status" or normalized == "staus":
                status_columns.append(column_idx)

        if status_columns:
            mapping["contract_status"] = status_columns[0]
            mapping["payment_status"] = status_columns[-1]
            if len(status_columns) > 1:
                mapping["extra_status"] = status_columns[-1]
        return mapping

    def _find_auxiliary_revenue_header(self, sheet) -> Optional[tuple[int, int, int, int, int]]:
        for row_idx in range(1, min((sheet.max_row or 0), 25) + 1):
            normalized_values = {
                column_idx: _normalize_text(sheet.cell(row_idx, column_idx).value)
                for column_idx in range(1, min((sheet.max_column or 0), 10) + 1)
            }
            date_columns = [column_idx for column_idx, value in normalized_values.items() if value == "data"]
            amount_columns = [column_idx for column_idx, value in normalized_values.items() if value == "valor"]
            status_columns = [
                column_idx
                for column_idx, value in normalized_values.items()
                if value in {"status", "staus"}
            ]
            if date_columns and amount_columns and status_columns:
                amount_col = amount_columns[-1]
                if amount_col < 6:
                    continue
                desc_col = max(1, amount_col - 2)
                return (desc_col, date_columns[-1], amount_col, status_columns[-1], row_idx)
        return None

    def _parse_permuta_snapshot(self, sheet) -> Optional[FinancialSectionSnapshot]:
        header_row = None
        for row_idx in range(1, min((sheet.max_row or 0), 25) + 1):
            if (
                _normalize_text(sheet.cell(row_idx, 1).value) == "cliente"
                and _normalize_text(sheet.cell(row_idx, 2).value) == "permuta"
                and _normalize_text(sheet.cell(row_idx, 3).value) == "saldo"
            ):
                header_row = row_idx
                break
        if header_row is None:
            return None

        entries: list[FinancialEntry] = []
        for row_idx in range(header_row + 1, min((sheet.max_row or 0), 25) + 1):
            client = _clean_value(sheet.cell(row_idx, 1).value)
            permuta_term = _clean_value(sheet.cell(row_idx, 2).value)
            amount = _as_number(sheet.cell(row_idx, 3).value)
            if not client and amount is None:
                continue
            if _normalize_text(client) in {"subtotal", "total"}:
                continue
            if not client:
                continue
            entries.append(
                FinancialEntry(
                    entry_type="permuta",
                    sheet_name=sheet.title,
                    description=client,
                    amount=amount,
                    counterparty=client,
                    notes=permuta_term,
                )
            )
        if not entries:
            return None
        return FinancialSectionSnapshot(
            section_key="permuta",
            title="Permutas e creditos",
            total_amount=_sum_amount(entry.amount for entry in entries),
            entry_count=len(entries),
            entries=entries,
        )

    def _parse_debt_snapshot(self, sheet) -> Optional[FinancialSectionSnapshot]:
        entries: list[FinancialEntry] = []
        for row_idx in range(1, min(self._row_limit(sheet), 45) + 1):
            label = _clean_value(sheet.cell(row_idx, 2).value)
            amount = _as_number(sheet.cell(row_idx, 3).value)
            if not label and amount is None:
                continue
            normalized = _normalize_text(label)
            if normalized in {"divida", "total"}:
                continue
            if "divida" in normalized and amount is None:
                continue
            if not label:
                continue
            if amount is None and normalized not in {"a_pagar"}:
                continue
            entries.append(
                FinancialEntry(
                    entry_type="debt",
                    sheet_name=sheet.title,
                    description=label,
                    amount=amount,
                    status="Aberto" if normalized == "a_pagar" or (amount or 0) > 0 else None,
                    counterparty=label,
                )
            )
        if not entries:
            return None
        return FinancialSectionSnapshot(
            section_key="debt",
            title="Dividas e compromissos",
            total_amount=self._resolve_debt_outstanding(
                FinancialSectionSnapshot(
                    section_key="debt",
                    title="Dividas e compromissos",
                    entry_count=len(entries),
                    entries=entries,
                )
            ),
            entry_count=len(entries),
            entries=entries,
        )

    def _parse_financial_sections(self, sheet) -> list[FinancialSectionSnapshot]:
        snapshots: dict[tuple[str, str], FinancialSectionSnapshot] = {}
        row_limit = self._row_limit(sheet)
        for label_col, value_col, date_col, status_col in ((5, 6, 7, 8), (10, 11, 12, 13)):
            current_owner: Optional[str] = None
            current_section: Optional[tuple[str, str]] = None
            for row_idx in range(1, row_limit + 1):
                label = _clean_value(sheet.cell(row_idx, label_col).value)
                amount = _as_number(sheet.cell(row_idx, value_col).value)
                date_value = _stringify_date(sheet.cell(row_idx, date_col).value)
                status = _clean_value(sheet.cell(row_idx, status_col).value)
                if not label and amount is None and not date_value and not status:
                    continue

                normalized = _normalize_text(label)
                if normalized in SECTION_TYPE_MAP:
                    section_key, section_title = SECTION_TYPE_MAP[normalized]
                    owner = current_owner or "Consolidado"
                    current_section = (owner, section_key)
                    snapshots.setdefault(
                        current_section,
                        FinancialSectionSnapshot(
                            section_key=section_key,
                            title=section_title,
                            owner_label=owner,
                        ),
                    )
                    continue

                if current_section is None and self._looks_like_owner_label(label):
                    current_owner = label
                    continue

                if current_section is None:
                    continue

                snapshot = snapshots[current_section]
                if normalized in {"total", "subtotal"}:
                    if amount is not None:
                        snapshot.total_amount = amount
                    current_section = None
                    continue

                if normalized.startswith("totalizacao") or any(
                    token in normalized for token in ("despesas_globais", "resultado_final", "recebiveis_modulo_versa")
                ):
                    current_section = None
                    continue

                if self._is_section_header_row(label, amount, date_value, status):
                    continue

                if not label:
                    continue

                snapshot.entries.append(
                    FinancialEntry(
                        entry_type=snapshot.section_key,
                        sheet_name=sheet.title,
                        description=label,
                        amount=amount,
                        status=status,
                        date=date_value,
                        due_date=date_value if snapshot.section_key in {"fixed_cost", "operating_cost"} else None,
                        counterparty=label,
                        owner_label=snapshot.owner_label,
                    )
                )
                snapshot.entry_count += 1

        for snapshot in snapshots.values():
            if snapshot.total_amount is None:
                snapshot.total_amount = _sum_amount(entry.amount for entry in snapshot.entries)

        return list(snapshots.values())

    def _extract_summary_values(self, sheet) -> dict[str, float]:
        summary: dict[str, float] = {}
        for row_idx in range(1, self._row_limit(sheet) + 1):
            for column_idx in range(1, min((sheet.max_column or 0), 13) + 1):
                label = _clean_value(sheet.cell(row_idx, column_idx).value)
                if not label:
                    continue
                normalized = _normalize_text(label)
                field_name = None
                for token, candidate in SUMMARY_FIELD_MAP.items():
                    if token in normalized:
                        field_name = candidate
                        break
                if field_name is None and normalized == "total" and row_idx >= max(20, (sheet.max_row or 0) - 15):
                    field_name = "closing_total"
                if field_name is None:
                    continue
                amount = self._find_number_to_right(sheet, row_idx, column_idx)
                if amount is None:
                    continue
                summary[field_name] = amount
        return summary

    def _find_number_to_right(self, sheet, row_idx: int, column_idx: int) -> Optional[float]:
        for offset in range(1, 5):
            amount = _as_number(sheet.cell(row_idx, column_idx + offset).value)
            if amount is not None:
                return amount
        return None

    def _build_financial_analysis(
        self,
        entity_name: Optional[str],
        months: list[FinancialPeriodSummary],
        summary_notes: list[str],
        detected_entities: set[str],
    ) -> FinancialAnalysisResult:
        sorted_months = sorted(months, key=self._period_sort_key)
        client_rollups, client_period_rollups, contract_rollups = self._build_receivable_rollups(sorted_months)
        revenue_total = _sum_amount(
            self._normalize_financial_amount(self._period_revenue(month))
            for month in sorted_months
        ) or 0.0
        gross_total = _sum_amount(
            self._normalize_financial_amount(month.gross_revenue_total)
            for month in sorted_months
        ) or revenue_total
        other_income_total = _sum_amount(
            self._normalize_financial_amount(month.other_income_total)
            for month in sorted_months
        ) or 0.0
        permuta_total = _sum_amount(
            self._normalize_financial_amount(month.permuta_balance)
            for month in sorted_months
        ) or 0.0
        taxes_total = _sum_amount(
            self._normalize_financial_amount(month.taxes_total)
            for month in sorted_months
        ) or 0.0
        personnel_total = _sum_amount(
            self._normalize_financial_amount(month.personnel_total)
            for month in sorted_months
        ) or 0.0
        fixed_costs_total = _sum_amount(
            self._normalize_financial_amount(month.fixed_costs_total)
            for month in sorted_months
        ) or 0.0
        operating_costs_total = _sum_amount(
            self._normalize_financial_amount(month.operating_costs_total)
            for month in sorted_months
        ) or 0.0
        global_expenses_total = (
            _sum_amount(
                self._normalize_financial_amount(month.global_expenses_total)
                for month in sorted_months
            )
            or taxes_total + personnel_total + fixed_costs_total + operating_costs_total
        )
        operating_inflows_total = revenue_total + other_income_total
        open_receivables_total = _sum_amount(
            self._period_open_receivables(month)
            for month in sorted_months
        )
        reported_net_result_total = _sum_amount(month.net_result for month in sorted_months)
        estimated_net_result_total = operating_inflows_total - global_expenses_total
        if permuta_total:
            estimated_net_result_total -= permuta_total
        net_result_total = (
            reported_net_result_total
            if reported_net_result_total is not None
            else estimated_net_result_total
        )
        latest_closing = next(
            (month.closing_total for month in reversed(sorted_months) if month.closing_total is not None),
            None,
        )
        lines = [
            self._build_dre_line(
                key="gross_revenue",
                label="Receita operacional bruta",
                amount=gross_total,
                line_type="revenue",
                gross_base=gross_total,
                operating_base=operating_inflows_total,
            ),
            self._build_dre_line(
                key="receivables",
                label="Receita base do fechamento",
                amount=revenue_total,
                line_type="revenue",
                gross_base=gross_total,
                operating_base=operating_inflows_total,
            ),
            self._build_dre_line(
                key="other_income",
                label="Outras entradas operacionais",
                amount=other_income_total,
                line_type="revenue",
                gross_base=gross_total,
                operating_base=operating_inflows_total,
            ),
            self._build_dre_line(
                key="operating_inflows",
                label="Entradas operacionais consideradas",
                amount=operating_inflows_total,
                line_type="revenue",
                gross_base=gross_total,
                operating_base=operating_inflows_total,
            ),
            self._build_dre_line(
                key="permuta",
                label="Permutas e creditos correlatos",
                amount=permuta_total,
                line_type="deduction",
                gross_base=gross_total,
                operating_base=operating_inflows_total,
            ),
            self._build_dre_line(
                key="taxes",
                label="Tributos e encargos",
                amount=taxes_total,
                line_type="expense",
                gross_base=gross_total,
                operating_base=operating_inflows_total,
            ),
            self._build_dre_line(
                key="personnel",
                label="Despesas com pessoal",
                amount=personnel_total,
                line_type="expense",
                gross_base=gross_total,
                operating_base=operating_inflows_total,
            ),
            self._build_dre_line(
                key="fixed_costs",
                label="Custos fixos",
                amount=fixed_costs_total,
                line_type="expense",
                gross_base=gross_total,
                operating_base=operating_inflows_total,
            ),
            self._build_dre_line(
                key="operating_costs",
                label="Custos operacionais",
                amount=operating_costs_total,
                line_type="expense",
                gross_base=gross_total,
                operating_base=operating_inflows_total,
            ),
            self._build_dre_line(
                key="global_expenses",
                label="Total de custos e despesas",
                amount=global_expenses_total,
                line_type="expense",
                gross_base=gross_total,
                operating_base=operating_inflows_total,
            ),
        ]
        if open_receivables_total is not None:
            lines.append(
                self._build_dre_line(
                    key="open_receivables",
                    label="Carteira em aberto",
                    amount=open_receivables_total,
                    line_type="balance",
                    gross_base=gross_total,
                    operating_base=operating_inflows_total,
                )
            )
        lines.append(
            self._build_dre_line(
                key="net_result",
                label="Resultado consolidado do periodo",
                amount=net_result_total,
                line_type="result",
                gross_base=gross_total,
                operating_base=operating_inflows_total,
            )
        )
        if latest_closing is not None:
            lines.append(
                self._build_dre_line(
                    key="closing_balance",
                    label="Saldo acumulado final informado",
                    amount=latest_closing,
                    line_type="balance",
                    gross_base=gross_total,
                    operating_base=operating_inflows_total,
                )
            )

        fiscal_year = next((month.year for month in sorted_months if month.year is not None), None)
        cleaned_entities = sorted(
            entity
            for entity in detected_entities
            if self._is_meaningful_detected_entity(entity, entity_name)
        )

        return FinancialAnalysisResult(
            fiscal_year=fiscal_year,
            entity_name=entity_name,
            months=sorted_months,
            dre_lines=lines,
            client_rollups=client_rollups,
            client_period_rollups=client_period_rollups,
            contract_rollups=contract_rollups,
            summary_notes=self._deduplicate_notes(summary_notes),
            detected_entities=cleaned_entities,
            entry_count=sum(
                section.entry_count
                for month in sorted_months
                for section in month.sections
            ),
        )

    def _build_receivable_rollups(
        self,
        months: list[FinancialPeriodSummary],
    ) -> tuple[list[FinancialClientRollup], list[FinancialClientPeriodRollup], list[FinancialContractRollup]]:
        client_buckets: dict[str, dict[str, object]] = {}
        client_period_buckets: dict[str, dict[str, object]] = {}
        contract_buckets: dict[str, dict[str, object]] = {}
        period_order = {month.period_label: index for index, month in enumerate(months)}

        for month in months:
            for section in month.sections:
                for entry in section.entries:
                    if entry.entry_type != "receivable":
                        continue
                    amount = self._normalize_financial_amount(entry.amount)
                    if amount is None:
                        continue

                    client_name = self._resolve_financial_client_name(entry)
                    contract_label = entry.contract_label or self._fallback_contract_label(entry)
                    is_realized = self._is_realized_receivable_entry(entry, section.section_key)

                    if client_name:
                        client_key = _normalize_text(client_name)
                        bucket = client_buckets.setdefault(
                            client_key,
                            {
                                "display_name": client_name,
                                "received": 0.0,
                                "expected": 0.0,
                                "pending": 0.0,
                                "months": [],
                                "contracts": [],
                            },
                        )
                        bucket["display_name"] = self._prefer_display_text(bucket["display_name"], client_name)
                        bucket["expected"] = float(bucket["expected"]) + amount
                        if is_realized:
                            bucket["received"] = float(bucket["received"]) + amount
                        else:
                            bucket["pending"] = float(bucket["pending"]) + amount
                        self._append_unique_text(bucket["months"], month.period_label)
                        if contract_label:
                            self._append_unique_text(bucket["contracts"], contract_label)

                        client_period_key = f"{client_key}::{month.period_label}"
                        period_bucket = client_period_buckets.setdefault(
                            client_period_key,
                            {
                                "display_name": client_name,
                                "period_label": month.period_label,
                                "received": 0.0,
                                "expected": 0.0,
                                "pending": 0.0,
                                "contracts": [],
                            },
                        )
                        period_bucket["display_name"] = self._prefer_display_text(
                            period_bucket["display_name"],
                            client_name,
                        )
                        period_bucket["expected"] = float(period_bucket["expected"]) + amount
                        if is_realized:
                            period_bucket["received"] = float(period_bucket["received"]) + amount
                        else:
                            period_bucket["pending"] = float(period_bucket["pending"]) + amount
                        if contract_label:
                            self._append_unique_text(period_bucket["contracts"], contract_label)

                    if not contract_label:
                        continue

                    contract_key = _normalize_text(contract_label)
                    contract_bucket = contract_buckets.setdefault(
                        contract_key,
                        {
                            "contract_label": contract_label,
                            "client_name": client_name,
                            "unit": entry.unit,
                            "contract_start_date": entry.contract_start_date,
                            "contract_end_date": entry.contract_end_date,
                            "latest_status": entry.status,
                            "received": 0.0,
                            "expected": 0.0,
                            "pending": 0.0,
                            "entry_count": 0,
                            "months": [],
                            "sheet_names": [],
                        },
                    )
                    contract_bucket["contract_label"] = self._prefer_display_text(
                        contract_bucket["contract_label"],
                        contract_label,
                    )
                    contract_bucket["client_name"] = self._prefer_display_text(
                        contract_bucket["client_name"],
                        client_name,
                    )
                    contract_bucket["unit"] = contract_bucket["unit"] or entry.unit
                    contract_bucket["contract_start_date"] = (
                        contract_bucket["contract_start_date"] or entry.contract_start_date
                    )
                    contract_bucket["contract_end_date"] = (
                        contract_bucket["contract_end_date"] or entry.contract_end_date
                    )
                    if entry.status:
                        contract_bucket["latest_status"] = entry.status
                    contract_bucket["expected"] = float(contract_bucket["expected"]) + amount
                    if is_realized:
                        contract_bucket["received"] = float(contract_bucket["received"]) + amount
                    else:
                        contract_bucket["pending"] = float(contract_bucket["pending"]) + amount
                    contract_bucket["entry_count"] = int(contract_bucket["entry_count"]) + 1
                    self._append_unique_text(contract_bucket["months"], month.period_label)
                    self._append_unique_text(contract_bucket["sheet_names"], entry.sheet_name)

        client_rollups = sorted(
            (
                FinancialClientRollup(
                    client_name=str(bucket["display_name"]),
                    total_received_amount=float(bucket["received"]),
                    total_expected_amount=float(bucket["expected"]),
                    total_pending_amount=float(bucket["pending"]),
                    contract_count=len(bucket["contracts"]),
                    months_covered=list(bucket["months"]),
                    contract_labels=list(bucket["contracts"]),
                )
                for bucket in client_buckets.values()
            ),
            key=lambda item: (-item.total_received_amount, -item.total_expected_amount, item.client_name.lower()),
        )

        client_period_rollups = sorted(
            (
                FinancialClientPeriodRollup(
                    client_name=str(bucket["display_name"]),
                    period_label=str(bucket["period_label"]),
                    total_received_amount=float(bucket["received"]),
                    total_expected_amount=float(bucket["expected"]),
                    total_pending_amount=float(bucket["pending"]),
                    contract_count=len(bucket["contracts"]),
                    contract_labels=list(bucket["contracts"]),
                )
                for bucket in client_period_buckets.values()
            ),
            key=lambda item: (
                period_order.get(item.period_label, 10**9),
                -item.total_received_amount,
                -item.total_expected_amount,
                item.client_name.lower(),
            ),
        )

        contract_rollups = sorted(
            (
                FinancialContractRollup(
                    contract_label=str(bucket["contract_label"]),
                    client_name=bucket["client_name"],
                    unit=bucket["unit"],
                    contract_start_date=bucket["contract_start_date"],
                    contract_end_date=bucket["contract_end_date"],
                    latest_status=bucket["latest_status"],
                    total_received_amount=float(bucket["received"]),
                    total_expected_amount=float(bucket["expected"]),
                    total_pending_amount=float(bucket["pending"]),
                    entry_count=int(bucket["entry_count"]),
                    months_covered=list(bucket["months"]),
                    source_sheet_names=list(bucket["sheet_names"]),
                )
                for bucket in contract_buckets.values()
            ),
            key=lambda item: (-item.total_received_amount, -item.total_expected_amount, item.contract_label.lower()),
        )
        return client_rollups, client_period_rollups, contract_rollups

    def _resolve_financial_client_name(self, entry: FinancialEntry) -> Optional[str]:
        return _clean_value(entry.counterparty) or _clean_value(entry.owner_label) or _truncate_text(entry.description, 120)

    def _fallback_contract_label(self, entry: FinancialEntry) -> Optional[str]:
        if not any((entry.unit, entry.contract_start_date, entry.contract_end_date)):
            return None
        return self._build_contract_label(
            client_name=self._resolve_financial_client_name(entry),
            unit=entry.unit,
            contract_start_date=entry.contract_start_date,
            contract_end_date=entry.contract_end_date,
            term=None,
        )

    def _is_realized_receivable_entry(self, entry: FinancialEntry, section_key: str) -> bool:
        if section_key == "receivable_open":
            return False
        normalized_status = _normalize_text(entry.status)
        if not normalized_status:
            return False
        if _is_pending_status(entry.status):
            return False
        return any(token in normalized_status for token in REALIZED_RECEIVABLE_STATUSES)

    def _build_contract_label(
        self,
        client_name: Optional[str],
        unit: Optional[str],
        contract_start_date: Optional[str],
        contract_end_date: Optional[str],
        term: Optional[str],
    ) -> Optional[str]:
        client = _clean_value(client_name)
        if not client:
            return None

        parts = [client]
        unit_text = _clean_value(unit)
        if unit_text:
            normalized_unit = _normalize_text(unit_text)
            parts.append(unit_text if normalized_unit.startswith("unidade") else f"Unidade {unit_text}")

        contract_anchor = (
            self._humanize_contract_marker(contract_start_date, "inicio")
            or _clean_value(term)
            or self._humanize_contract_marker(contract_end_date, "fim")
        )
        if contract_anchor:
            parts.append(contract_anchor)
        return " | ".join(parts)

    def _humanize_contract_marker(self, value: Optional[str], prefix: str) -> Optional[str]:
        text = _clean_value(value)
        if not text:
            return None
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
            year, month, day = text.split("-")
            return f"{prefix} {day}/{month}/{year}"
        return text

    def _append_unique_text(self, target: object, value: Optional[str]) -> None:
        if not isinstance(target, list):
            return
        text = _clean_value(value)
        if not text or text in target:
            return
        target.append(text)

    def _prefer_display_text(self, current: Optional[str], candidate: Optional[str]) -> Optional[str]:
        current_text = _clean_value(current)
        candidate_text = _clean_value(candidate)
        if not candidate_text:
            return current_text
        if not current_text:
            return candidate_text
        if len(candidate_text) > len(current_text):
            return candidate_text
        return current_text

    def _build_dre_line(
        self,
        key: str,
        label: str,
        amount: Optional[float],
        line_type: str,
        gross_base: Optional[float],
        operating_base: Optional[float],
    ) -> FinancialStatementLine:
        resolved_amount = float(amount or 0.0)
        return FinancialStatementLine(
            key=key,
            label=label,
            amount=resolved_amount,
            line_type=line_type,
            share_of_gross_revenue=self._line_share(resolved_amount, gross_base),
            share_of_operating_inflows=self._line_share(resolved_amount, operating_base),
        )

    def _build_context_layers(self, analysis: FinancialAnalysisResult) -> list[WorkbookContextLayer]:
        layers: list[WorkbookContextLayer] = [
            WorkbookContextLayer(
                layer_type="financial_results_bridge",
                sheet_name="Consolidado",
                title="Painel consolidado da DRE",
                summary=(
                    f"{len(analysis.months)} periodo(s) consolidado(s), "
                    f"entradas operacionais de {_format_currency(self._line_amount(analysis, 'operating_inflows'))}, "
                    f"custos e despesas de {_format_currency(self._line_amount(analysis, 'global_expenses'))} "
                    f"e resultado consolidado de {_format_currency(self._line_amount(analysis, 'net_result'))}."
                ),
                details=[
                    f"{line.label}: {_format_currency(line.amount)}"
                    for line in analysis.dre_lines[:7]
                ],
            )
        ]
        if analysis.client_rollups:
            top_clients = analysis.client_rollups[:6]
            layers.append(
                WorkbookContextLayer(
                    layer_type="financial_client_rollup",
                    sheet_name="Consolidado",
                    title="Recebimentos por cliente",
                    summary=(
                        f"{len(analysis.client_rollups)} cliente(s) consolidado(s), com destaque para "
                        + " | ".join(
                            f"{client.client_name}: {_format_currency(client.total_received_amount)}"
                            for client in top_clients[:3]
                        )
                        + "."
                    ),
                    details=[
                        (
                            f"{client.client_name}: recebido {_format_currency(client.total_received_amount)}, "
                            f"pendente {_format_currency(client.total_pending_amount)}, "
                            f"contratos {client.contract_count}"
                        )
                        for client in top_clients
                    ],
                )
            )
        if analysis.contract_rollups:
            top_contracts = analysis.contract_rollups[:6]
            layers.append(
                WorkbookContextLayer(
                    layer_type="financial_contract_rollup",
                    sheet_name="Consolidado",
                    title="Recebimentos por contrato",
                    summary=(
                        f"{len(analysis.contract_rollups)} contrato(s) estruturado(s), com maior rendimento em "
                        + " | ".join(
                            f"{contract.contract_label}: {_format_currency(contract.total_received_amount)}"
                            for contract in top_contracts[:3]
                        )
                        + "."
                    ),
                    details=[
                        (
                            f"{contract.contract_label}: recebido {_format_currency(contract.total_received_amount)}, "
                            f"previsto {_format_currency(contract.total_expected_amount)}, "
                            f"pendente {_format_currency(contract.total_pending_amount)}"
                        )
                        for contract in top_contracts
                    ],
                )
            )
        for month in analysis.months:
            layers.append(
                WorkbookContextLayer(
                    layer_type="financial_overview",
                    sheet_name=month.sheet_name,
                    title=f"Fechamento mensal de {month.period_label}",
                    summary=(
                        f"Receita base de {_format_currency(self._normalize_financial_amount(month.receivables_total))}, "
                        f"custos e despesas de {_format_currency(self._normalize_financial_amount(month.global_expenses_total))} "
                        f"e resultado de {_format_currency(month.net_result)}."
                    ),
                    details=[
                        f"Pendencias abertas: {month.pending_entry_count}",
                        f"Impostos: {_format_currency(self._normalize_financial_amount(month.taxes_total))}",
                        f"Pessoal: {_format_currency(self._normalize_financial_amount(month.personnel_total))}",
                        f"Custos fixos: {_format_currency(self._normalize_financial_amount(month.fixed_costs_total))}",
                        f"Custos operacionais: {_format_currency(self._normalize_financial_amount(month.operating_costs_total))}",
                    ],
                )
            )
            cost_details = self._top_section_details(month)
            if cost_details:
                layers.append(
                    WorkbookContextLayer(
                        layer_type="financial_cost_structure",
                        sheet_name=month.sheet_name,
                        title=f"Composicao de custos de {month.period_label}",
                        summary=(
                            "Blocos com maior peso no periodo: " + " | ".join(cost_details[:3])
                        ),
                        details=cost_details,
                    )
                )
        return layers

    def _build_financial_warnings(self, analysis: FinancialAnalysisResult) -> list[str]:
        warnings: list[str] = []
        if not analysis.months:
            return warnings
        negative_months = [month for month in analysis.months if (month.net_result or 0) < 0]
        if negative_months:
            warnings.append(
                "Periodos com resultado negativo: "
                + ", ".join(
                    f"{month.period_label} ({_format_currency(month.net_result)})"
                    for month in negative_months[:6]
                )
                + "."
            )
        pending_months = [month for month in analysis.months if month.pending_entry_count > 0]
        if pending_months:
            warnings.append(
                "Ha pendencias operacionais registradas em "
                + ", ".join(
                    f"{month.period_label} ({month.pending_entry_count})"
                    for month in pending_months[:6]
                )
                + "."
            )
        total_inflows = self._line_amount(analysis, "operating_inflows") or 0.0
        total_expenses = self._line_amount(analysis, "global_expenses") or 0.0
        open_balance = self._line_amount(analysis, "open_receivables")
        if total_inflows and total_expenses and total_expenses / total_inflows >= 0.8:
            warnings.append(
                f"A pressao de custos e despesas atingiu {round((total_expenses / total_inflows) * 100)}% das entradas operacionais consolidadas."
            )
        if total_inflows and open_balance and open_balance / total_inflows >= 0.3:
            warnings.append(
                f"A carteira em aberto representa {round((open_balance / total_inflows) * 100)}% das entradas operacionais consideradas no demonstrativo."
            )
        return warnings

    def _build_period_notes(
        self,
        sheet_name: str,
        revenue_block: RevenueBlockResult,
        summary_values: dict[str, float],
        section_totals: dict[str, Optional[float]],
    ) -> list[str]:
        notes: list[str] = []
        if revenue_block.other_income_entries:
            notes.append(
                f"{sheet_name}: foram identificadas {len(revenue_block.other_income_entries)} entrada(s) complementar(es)."
            )
        if summary_values.get("closing_total") is not None:
            notes.append(
                f"{sheet_name}: saldo acumulado informado em {_format_currency(summary_values.get('closing_total'))}."
            )
        top_section = max(
            (
                (section_key, abs(amount))
                for section_key, amount in section_totals.items()
                if amount is not None
            ),
            default=None,
            key=lambda item: item[1],
        )
        if top_section is not None:
            notes.append(
                f"{sheet_name}: maior bloco de custo no periodo = {top_section[0]} ({_format_currency(section_totals[top_section[0]])})."
            )
        return notes

    def _aggregate_section_totals(
        self,
        section_snapshots: list[FinancialSectionSnapshot],
    ) -> dict[str, Optional[float]]:
        totals: dict[str, Optional[float]] = defaultdict(lambda: None)
        for snapshot in section_snapshots:
            current = totals.get(snapshot.section_key)
            if snapshot.total_amount is None:
                continue
            totals[snapshot.section_key] = _first_number(current, 0.0) + snapshot.total_amount
        return dict(totals)

    def _resolve_debt_outstanding(self, debt_snapshot: Optional[FinancialSectionSnapshot]) -> Optional[float]:
        if debt_snapshot is None:
            return None
        for entry in debt_snapshot.entries:
            if _normalize_text(entry.description) == "a_pagar" and entry.amount is not None:
                return entry.amount
        return _sum_amount(
            entry.amount
            for entry in debt_snapshot.entries
            if entry.amount is not None and entry.amount > 0
        )

    def _top_section_details(self, month: FinancialPeriodSummary) -> list[str]:
        ranked = sorted(
            (
                snapshot
                for snapshot in month.sections
                if snapshot.total_amount is not None and snapshot.section_key in {"tax", "personnel", "fixed_cost", "operating_cost"}
            ),
            key=lambda snapshot: abs(snapshot.total_amount or 0.0),
            reverse=True,
        )
        return [
            f"{snapshot.title} ({snapshot.owner_label or 'Consolidado'}): {_format_currency(self._normalize_financial_amount(snapshot.total_amount))}"
            for snapshot in ranked[:6]
        ]

    def _period_open_receivables(self, period: FinancialPeriodSummary) -> Optional[float]:
        for section in period.sections:
            if section.section_key != "receivable_open":
                continue
            total_amount = section.total_amount
            if total_amount is None:
                total_amount = _sum_amount(entry.amount for entry in section.entries)
            return self._normalize_financial_amount(total_amount)
        if period.debt_outstanding is not None:
            return self._normalize_financial_amount(period.debt_outstanding)
        return None

    def _extract_reference_links(
        self,
        workbook,
        selected_sheet_names: list[str],
    ) -> list[WorkbookReferenceLink]:
        links: list[WorkbookReferenceLink] = []
        seen_urls: set[str] = set()
        selected_sheets = {_normalize_text(name) for name in selected_sheet_names}
        for sheet in workbook.worksheets:
            normalized_sheet = _normalize_text(sheet.title)
            if selected_sheets and normalized_sheet not in selected_sheets:
                continue
            for row in sheet.iter_rows(max_row=self._row_limit(sheet)):
                row_values = [
                    (cell.coordinate, _clean_value(cell.value))
                    for cell in row
                    if _clean_value(cell.value)
                ]
                if not row_values:
                    continue
                row_text = " | ".join(value for _, value in row_values)
                for coordinate, value in row_values:
                    for url in _extract_urls(value):
                        normalized_url = _normalize_url_for_dedup(url)
                        if not normalized_url or normalized_url in seen_urls:
                            continue
                        seen_urls.add(normalized_url)
                        links.append(
                            WorkbookReferenceLink(
                                url=url,
                                sheet_name=sheet.title,
                                cell_reference=coordinate,
                                label=self._reference_label(row_values, coordinate, url),
                                context=_truncate_text(row_text, 220),
                                source_hint="nao_informada",
                                link_kind="reference",
                                crawlable=_is_crawlable_reference_url(url),
                            )
                        )
        return links

    def _reference_label(
        self,
        row_values: list[tuple[str, str]],
        coordinate: str,
        url: str,
    ) -> Optional[str]:
        for cell_coordinate, value in row_values:
            if cell_coordinate == coordinate or url in value:
                continue
            return _truncate_text(value, 120)
        return None

    def _guess_entity_from_periods(
        self,
        periods: list[FinancialPeriodSummary],
        filename_hint: Path,
    ) -> Optional[str]:
        filename_entity = self._infer_entity_name(filename_hint)
        if filename_entity:
            return filename_entity
        for month in periods:
            for section in month.sections:
                if section.owner_label and "versa" in _normalize_text(section.owner_label):
                    return section.owner_label
        return None

    def _build_period_label(self, sheets: list, filename_hint: Path) -> Optional[str]:
        if not sheets:
            return None
        year = self._detect_year_from_path(filename_hint)
        month_labels = [self._month_label(sheet.title) for sheet in sheets]
        if not month_labels:
            return str(year) if year else None
        unique_labels = list(dict.fromkeys(month_labels))
        if len(unique_labels) == 1:
            return f"{unique_labels[0]}/{year}" if year else unique_labels[0]
        return f"{unique_labels[0]} a {unique_labels[-1]}/{year}" if year else f"{unique_labels[0]} a {unique_labels[-1]}"

    def _infer_entity_name(self, filename_hint: Path) -> Optional[str]:
        stem = _clean_value(filename_hint.stem)
        if not stem:
            return None
        cleaned = re.sub(r"\b(20\d{2}|19\d{2})\b", "", stem)
        cleaned = re.sub(
            r"\b(planilha|controle|geral|custo|custo[s]?|novo|de|do|da|dos|das|financeiro|financeira|contabil|inicio|vbc)\b",
            "",
            cleaned,
            flags=re.IGNORECASE,
        )
        cleaned = re.sub(r"[-_]+", " ", cleaned)
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" -_")
        if not cleaned or len(cleaned) <= 3:
            return None
        return cleaned

    def _detect_sheet_year(self, sheet, filename_hint: Path) -> Optional[int]:
        filename_year = self._detect_year_from_path(filename_hint)
        if filename_year is not None:
            return filename_year
        for row_idx in range(1, min((sheet.max_row or 0), 20) + 1):
            for column_idx in range(1, min((sheet.max_column or 0), 12) + 1):
                value = sheet.cell(row_idx, column_idx).value
                if isinstance(value, datetime):
                    return value.year
                text = _clean_value(value)
                if not text:
                    continue
                match = re.search(r"\b(20\d{2}|19\d{2})\b", text)
                if match:
                    return int(match.group(1))
        return self._detect_year_from_path(filename_hint)

    def _detect_year_from_path(self, filename_hint: Path) -> Optional[int]:
        match = re.search(r"\b(20\d{2}|19\d{2})\b", filename_hint.stem)
        if match:
            return int(match.group(1))
        return None

    def _month_label(self, sheet_name: str) -> str:
        normalized = _normalize_text(self._base_sheet_name(sheet_name))
        month_data = MONTH_NAME_MAP.get(normalized)
        if month_data:
            return month_data[1]
        return sheet_name.strip()

    def _period_sort_key(self, period: FinancialPeriodSummary) -> tuple[int, int, str]:
        base_sheet_name = self._base_sheet_name(period.sheet_name)
        normalized = _normalize_text(base_sheet_name)
        month_data = MONTH_NAME_MAP.get(normalized, (99, base_sheet_name))
        return (period.year or 0, month_data[0], base_sheet_name.lower())

    def _base_sheet_name(self, sheet_name: str) -> str:
        return sheet_name.split("::", 1)[-1].strip() if "::" in sheet_name else sheet_name

    def _looks_like_owner_label(self, value: Optional[str]) -> bool:
        normalized = _normalize_text(value)
        return any(token in normalized for token in ("distribuidor", "versa"))

    def _is_section_header_row(
        self,
        label: Optional[str],
        amount: Optional[float],
        date_value: Optional[str],
        status: Optional[str],
    ) -> bool:
        normalized = _normalize_text(label)
        if normalized in {"valor", "data", "vencimento", "status", "staus"}:
            return True
        if amount is None and not date_value and not status and normalized in {"vbc_distribuidor", "modulo_versa"}:
            return True
        return False

    def _join_notes(self, *values: Optional[str]) -> Optional[str]:
        cleaned = [value for value in values if value]
        if not cleaned:
            return None
        return " | ".join(cleaned)

    def _row_values(self, sheet, row_idx: int, max_columns: int) -> dict[str, str]:
        values: dict[str, str] = {}
        for column_idx in range(1, min((sheet.max_column or 0), max_columns) + 1):
            cleaned = _clean_value(sheet.cell(row_idx, column_idx).value)
            if cleaned:
                values[str(column_idx)] = cleaned
        return values

    def _cell_text(self, sheet, row_idx: int, column_idx: Optional[int]) -> Optional[str]:
        if not column_idx:
            return None
        return _clean_value(sheet.cell(row_idx, column_idx).value)

    def _cell_date(self, sheet, row_idx: int, column_idx: Optional[int]) -> Optional[str]:
        if not column_idx:
            return None
        return _stringify_date(sheet.cell(row_idx, column_idx).value)

    def _iter_period_entries(self, period: FinancialPeriodSummary):
        for section in period.sections:
            for entry in section.entries:
                yield entry

    def _iter_section_entries(self, sections: list[FinancialSectionSnapshot]):
        for section in sections:
            for entry in section.entries:
                yield entry

    def _expand_month_entries(self, months: list[FinancialPeriodSummary]):
        for month in months:
            for section in month.sections:
                yield section

    def _line_amount(self, analysis: FinancialAnalysisResult, key: str) -> Optional[float]:
        for line in analysis.dre_lines:
            if line.key == key:
                return line.amount
        return None

    def _line_share(self, amount: Optional[float], base: Optional[float]) -> Optional[float]:
        if amount is None or base is None:
            return None
        resolved_base = float(base)
        if abs(resolved_base) < 0.0001:
            return None
        return abs(float(amount)) / abs(resolved_base)

    def _period_revenue(self, period: FinancialPeriodSummary) -> Optional[float]:
        return period.receivables_total if period.receivables_total is not None else period.gross_revenue_total

    def _normalize_financial_amount(self, value: Optional[float]) -> Optional[float]:
        if value is None:
            return None
        return abs(float(value))

    def _deduplicate_notes(self, notes: list[str]) -> list[str]:
        deduplicated: list[str] = []
        seen: set[str] = set()
        for note in notes:
            normalized = " ".join((note or "").split())
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            deduplicated.append(normalized)
        return deduplicated

    def _row_limit(self, sheet) -> int:
        cached = self._sheet_row_limits.get(sheet.title)
        if cached is not None:
            return cached

        upper_bound = min(int(sheet.max_row or 0), 2000)
        last_nonempty = 0
        blank_streak = 0
        for row_idx in range(1, upper_bound + 1):
            has_value = any(
                _clean_value(sheet.cell(row_idx, column_idx).value)
                for column_idx in range(1, min(int(sheet.max_column or 0), 14) + 1)
            )
            if has_value:
                last_nonempty = row_idx
                blank_streak = 0
                continue
            if last_nonempty:
                blank_streak += 1
                if blank_streak >= 40 and row_idx > 50:
                    break

        resolved = max(last_nonempty, 1)
        self._sheet_row_limits[sheet.title] = resolved
        return resolved

    def _is_meaningful_detected_entity(
        self,
        value: Optional[str],
        entity_name: Optional[str],
    ) -> bool:
        cleaned = _clean_value(value)
        if not cleaned:
            return False
        normalized = _normalize_text(cleaned)
        if not normalized or normalized == _normalize_text(entity_name):
            return False
        if normalized in {"-", "a_pagar", "total", "subtotal", "consolidado"}:
            return False
        if re.fullmatch(r"\d+_mes(?:es)?", normalized):
            return False
        if re.fullmatch(r"\d+_dias?", normalized):
            return False
        if normalized.startswith("aluguel_") and len(normalized) < 18:
            return False
        return True


def _normalize_text(value: object) -> str:
    cleaned = _clean_value(value)
    if not cleaned:
        return ""
    cleaned = unicodedata.normalize("NFKD", cleaned)
    cleaned = "".join(char for char in cleaned if not unicodedata.combining(char))
    cleaned = cleaned.lower()
    cleaned = cleaned.replace("-", "_").replace("/", "_")
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = re.sub(r"[^\w_]", "", cleaned)
    return cleaned.strip("_")


def _clean_value(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    if isinstance(value, datetime):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _as_number(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_value(value)
    if not text:
        return None
    normalized = text.replace("R$", "").replace(" ", "")
    if normalized in {"-", "x", "X"}:
        return None
    if normalized.count(",") == 1 and normalized.count(".") > 1:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif normalized.count(",") == 1 and normalized.count(".") == 0:
        normalized = normalized.replace(",", ".")
    else:
        normalized = normalized.replace(",", "")
    try:
        return float(normalized)
    except ValueError:
        return None


def _sum_amount(values) -> Optional[float]:
    numbers = [value for value in values if value is not None]
    if not numbers:
        return None
    return float(sum(numbers))


def _first_number(value: Optional[float], fallback: float) -> float:
    return value if value is not None else fallback


def _parse_sheet_names(value: Optional[str]) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in re.split(r"[,;\n]+", value) if part.strip()]


def _is_auto_sheet_selection(value: Optional[str]) -> bool:
    return _normalize_text(value) in AUTO_SHEET_SELECTIONS


def _extract_urls(value: str) -> list[str]:
    matches = re.findall(r"https?://\S+", value, re.IGNORECASE)
    cleaned: list[str] = []
    seen: set[str] = set()
    for match in matches:
        url = match.rstrip(").,;")
        if url in seen:
            continue
        seen.add(url)
        cleaned.append(url)
    return cleaned


def _normalize_url_for_dedup(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    parsed = urlparse(raw)
    host = (parsed.netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    path = (parsed.path or "").rstrip("/")
    query = f"?{parsed.query}" if parsed.query else ""
    return f"{parsed.scheme.lower()}://{host}{path}{query}".rstrip("/")


def _is_crawlable_reference_url(value: str) -> bool:
    lowered = (value or "").strip().lower()
    if not lowered.startswith(("http://", "https://")):
        return False
    return not any(lowered.endswith(extension) for extension in NON_CRAWLABLE_EXTENSIONS)


def _truncate_text(value: Optional[str], limit: int) -> Optional[str]:
    text = _clean_value(value)
    if not text:
        return None
    compact = " ".join(text.split())
    if len(compact) <= limit:
        return compact
    return compact[: limit - 3].rstrip() + "..."


def _stringify_date(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _clean_value(value)


def _is_pending_status(value: Optional[str]) -> bool:
    normalized = _normalize_text(value)
    if not normalized:
        return False
    return any(
        token in normalized
        for token in ("pendente", "parcial", "vencido", "atrasado", "a_vencer", "a_receber", "aberto")
    )


def _format_currency(value: Optional[float]) -> str:
    if value is None:
        return "-"
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
